"""Carregador de tabelas dinâmicas da ECF (RFB).

Mantém snapshot versionado das planilhas baixadas do Portal SPED,
conforme CLAUDE.md §2 e §23 do ecf-layout.md.

Caminho canônico: data/tabelas-dinamicas-rfb/ecf/ac-{YYYY}/Tabelas_Dinamicas_ECF_Leiaute_*_AC_{YYYY}_*.xlsx

Estrutura de cada tabela dinâmica: CÓDIGO | DESCRIÇÃO | DT_INI | DT_FIM | TIPO | ...
- TIPO 'R': rótulo/cabeçalho de seção (não é linha de dado)
- TIPO 'E': entrada de dado (valor preenchível pelo contribuinte)
- TIPO 'CNA' ou 'CA': cálculo automático ou alterável

O CÓDIGO é a chave de cruzamento com os campos CODIGO nos registros
M300, M350, X480, L300, N630, etc. da ECF.
"""

import logging
from functools import lru_cache
from pathlib import Path

logger = logging.getLogger(__name__)

# Mapeamento de registro ECF → nome da aba na planilha (variante A = PJ em Geral)
_ABA_POR_REGISTRO = {
    "M300": "M300A",
    "M350": "M350A",
    "X480": "X480",
    "L300": "L300A",
    "N630": "N630A",
    "N670": "N670A",
    "P150": "P150A",
    "PARTEB": "PARTEB_PADRAO",
}


def _caminho_planilha(ano_calendario: int) -> Path | None:
    """Retorna o caminho da planilha de tabelas dinâmicas para o ano-calendário."""
    base = Path(__file__).parent.parent.parent / "data" / "tabelas-dinamicas-rfb" / "ecf"
    pasta_ac = base / f"ac-{ano_calendario}"
    if not pasta_ac.exists():
        return None
    arquivos = list(pasta_ac.glob("Tabelas_Dinamicas_ECF_*.xlsx"))
    if not arquivos:
        return None
    return sorted(arquivos)[-1]  # mais recente se houver mais de um


@lru_cache(maxsize=8)
def carregar_tabela(registro: str, ano_calendario: int) -> dict[str, dict]:
    """Carrega a tabela dinâmica de um registro ECF para um dado ano-calendário.

    Returns:
        Dict {str(codigo): {'descricao': str, 'tipo': str, 'dt_ini': str}}
        Retorna dict vazio se a planilha não estiver disponível.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl não disponível — tabelas dinâmicas ECF não carregadas.")
        return {}

    caminho = _caminho_planilha(ano_calendario)
    if caminho is None:
        logger.warning(
            "Tabelas dinâmicas ECF para AC %d não encontradas em data/tabelas-dinamicas-rfb/ecf/ac-%d/. "
            "Campos CODIGO de registros %s serão armazenados sem decodificação semântica. "
            "CLAUDE.md §11: baixar e versionar antes de implementar decodificação.",
            ano_calendario, ano_calendario, registro,
        )
        return {}

    aba = _ABA_POR_REGISTRO.get(registro)
    if aba is None:
        return {}

    try:
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        if aba not in wb.sheetnames:
            logger.warning("Aba '%s' não encontrada em %s", aba, caminho.name)
            return {}

        ws = wb[aba]
        resultado: dict[str, dict] = {}
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if row[0] is None:
                continue
            codigo = str(row[0]).strip()
            descricao = str(row[1]).strip() if row[1] else ""
            dt_ini = str(row[2]).strip() if row[2] else ""
            tipo = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            resultado[codigo] = {
                "descricao": descricao,
                "tipo": tipo,
                "dt_ini": dt_ini,
            }
        logger.info(
            "Tabela dinâmica ECF '%s' AC %d carregada: %d entradas de %s",
            aba, ano_calendario, len(resultado), caminho.name,
        )
        return resultado
    except Exception as exc:
        logger.error("Erro ao carregar tabela dinâmica ECF '%s' AC %d: %s", aba, ano_calendario, exc)
        return {}


def tabela_disponivel(ano_calendario: int) -> bool:
    """Retorna True se o snapshot da planilha existe para o ano-calendário."""
    return _caminho_planilha(ano_calendario) is not None


def codigos_tipo_entrada(registro: str, ano_calendario: int) -> set[str]:
    """Retorna set de CODIGO com TIPO='E' (entradas de dado) da tabela dinâmica."""
    tabela = carregar_tabela(registro, ano_calendario)
    return {cod for cod, meta in tabela.items() if meta.get("tipo") == "E"}

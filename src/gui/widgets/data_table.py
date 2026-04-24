"""
DataTable — tabela densa com ordenação, filtro, seleção e exportação.

Componente central do design system (Bloco 5 §C1). Será o widget mais
usado da aplicação: T1 (clientes), T3 (matriz de cruzamentos), T4
(evidências), T6 (planos atual/antigo), T8 (auditoria).

Esta primeira iteração entrega a fundação: ColumnSpec dataclass + 4
column kinds (text, money, badge, code_link, int), filtro incremental
acoplado a SearchField, ordenação por header, seleção single-row,
exportação Excel da view atual, footer com contador e empty state.

Fora do escopo desta iteração (deferido para posterior):
  - Edição inline (decisão #28 — só quando integrarmos T6)
  - Multi-seleção range
  - Skeleton loading
  - Persistência de larguras de coluna em QSettings
  - Delegates: version, date, check
"""

from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable, Literal

from PySide6.QtCore import (
    QAbstractTableModel,
    QModelIndex,
    QSortFilterProxyModel,
    Qt,
    Signal,
)
from PySide6.QtGui import QColor, QFont, QIcon, QKeySequence, QPainter, QShortcut
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QPushButton,
    QStyle,
    QStyledItemDelegate,
    QStyleOptionViewItem,
    QTableView,
    QVBoxLayout,
    QWidget,
)

from src.gui.widgets.money_cell import Money, MoneyCellDelegate, MoneyOptions
from src.gui.widgets.search_field import SearchField
from src.gui.widgets.status_badge import _CORES as _BADGE_CORES, BadgeStatus


# --------------------------------------------------------------------
# ColumnSpec
# --------------------------------------------------------------------

ColumnKind = Literal["text", "money", "badge", "code_link", "int"]


@dataclass
class ColumnSpec:
    id: str                           # chave do dict da row
    header: str                       # texto do header
    kind: ColumnKind = "text"
    width: int | None = None
    align: Qt.AlignmentFlag = Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
    sortable: bool = True
    filterable: bool = True
    formatter: Callable[[Any], str] | None = None  # custom display
    money_options: MoneyOptions | None = None      # se kind="money"


# --------------------------------------------------------------------
# TableModel
# --------------------------------------------------------------------

class _TableModel(QAbstractTableModel):
    """Model que adapta `list[dict]` × `list[ColumnSpec]` para QTableView."""

    def __init__(self, columns: list[ColumnSpec], rows: list[dict], parent=None):
        super().__init__(parent)
        self._columns = list(columns)
        self._rows: list[dict] = list(rows)

    # -- API básica do QAbstractTableModel --------------------------

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:  # noqa: N802
        return 0 if parent.isValid() else len(self._rows)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:  # noqa: N802
        return 0 if parent.isValid() else len(self._columns)

    def headerData(  # noqa: N802
        self,
        section: int,
        orientation: Qt.Orientation,
        role: int = Qt.ItemDataRole.DisplayRole,
    ):
        if orientation == Qt.Orientation.Horizontal and role == Qt.ItemDataRole.DisplayRole:
            return self._columns[section].header
        if (
            orientation == Qt.Orientation.Horizontal
            and role == Qt.ItemDataRole.TextAlignmentRole
        ):
            return int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        return None

    def data(self, index: QModelIndex, role: int = Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        row = self._rows[index.row()]
        col = self._columns[index.column()]
        raw = row.get(col.id)

        if role == Qt.ItemDataRole.UserRole:
            # valor cru — usado pelo delegate e por sort.
            # Para colunas numéricas, converte Decimal → float para que o
            # QSortFilterProxyModel consiga ordenar (Qt não compara Decimal).
            if col.kind in ("money", "int") and raw is not None:
                try:
                    return float(Decimal(str(raw)))
                except (InvalidOperation, ValueError, TypeError):
                    return 0.0
            return raw

        if role == Qt.ItemDataRole.DisplayRole:
            return self._format_display(col, raw)

        if role == Qt.ItemDataRole.TextAlignmentRole:
            if col.kind in ("money", "int"):
                return int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            return int(col.align)

        if role == Qt.ItemDataRole.ToolTipRole:
            return self._tooltip(col, raw)

        return None

    # -- Helpers internos --------------------------------------------

    def _format_display(self, col: ColumnSpec, raw: Any) -> str:
        if col.formatter is not None:
            try:
                return col.formatter(raw)
            except Exception:
                return str(raw) if raw is not None else ""

        if raw is None:
            return ""

        if col.kind == "money":
            return Money.format(raw, col.money_options or MoneyOptions())

        if col.kind == "int":
            try:
                return f"{int(raw):,}".replace(",", ".")
            except (TypeError, ValueError):
                return str(raw)

        if col.kind == "badge":
            if isinstance(raw, BadgeStatus):
                return raw.value.upper()
            return str(raw).upper()

        if col.kind == "code_link":
            # Esperamos dict com 'linha' ou um número; renderiza "L{n}".
            if isinstance(raw, dict):
                return f"L{raw.get('linha', '?')}"
            return f"L{raw}"

        return str(raw)

    def _tooltip(self, col: ColumnSpec, raw: Any) -> str | None:
        if col.kind == "code_link" and isinstance(raw, dict):
            arq = Path(raw.get("arquivo", "")).name
            return (
                f"{arq}\nBloco {raw.get('bloco', '?')} · "
                f"Registro {raw.get('registro', '?')} · "
                f"Linha {raw.get('linha', '?')}"
            )
        if col.kind == "money":
            return Money.format_accessible(raw)
        return None

    # -- API de mutação ---------------------------------------------

    def set_rows(self, rows: list[dict]) -> None:
        self.beginResetModel()
        self._rows = list(rows)
        self.endResetModel()

    def add_row(self, row: dict) -> None:
        n = len(self._rows)
        self.beginInsertRows(QModelIndex(), n, n)
        self._rows.append(row)
        self.endInsertRows()

    def update_row(self, predicate: Callable[[dict], bool], patch: dict) -> int:
        """Aplica patch nas linhas que passam no predicate. Retorna quantidade alterada."""
        count = 0
        for i, r in enumerate(self._rows):
            if predicate(r):
                r.update(patch)
                top = self.index(i, 0)
                bot = self.index(i, len(self._columns) - 1)
                self.dataChanged.emit(top, bot)
                count += 1
        return count

    def remove_rows(self, predicate: Callable[[dict], bool]) -> int:
        ids = [i for i, r in enumerate(self._rows) if predicate(r)]
        if not ids:
            return 0
        for i in reversed(ids):
            self.beginRemoveRows(QModelIndex(), i, i)
            del self._rows[i]
            self.endRemoveRows()
        return len(ids)

    def rows(self) -> list[dict]:
        return list(self._rows)

    def columns(self) -> list[ColumnSpec]:
        return list(self._columns)


# --------------------------------------------------------------------
# Helper compartilhado pelos delegates: pinta só o fundo da célula
# --------------------------------------------------------------------

def _pintar_fundo_sem_texto(
    delegate: QStyledItemDelegate,
    painter: QPainter,
    option: QStyleOptionViewItem,
    index: QModelIndex,
) -> None:
    """Pinta seleção/hover/alternate-row sem desenhar o texto do DisplayRole.

    Necessário para delegates que vão pintar texto custom — caso contrário
    o estilo do Qt 6 desenha o DisplayRole por baixo do nosso paint, gerando
    duplicação visual ("texto borrado").
    """
    opt = QStyleOptionViewItem(option)
    delegate.initStyleOption(opt, index)
    opt.text = ""
    opt.icon = QIcon()
    opt.features &= ~QStyleOptionViewItem.ViewItemFeature.HasDecoration
    widget = option.widget
    style = widget.style() if widget is not None else QApplication.style()
    style.drawControl(QStyle.ControlElement.CE_ItemViewItem, opt, painter, widget)


# --------------------------------------------------------------------
# Delegate de Badge (pinta StatusBadge sem instanciar widget por célula)
# --------------------------------------------------------------------

class _BadgeCellDelegate(QStyledItemDelegate):
    """Pinta um pill estilo StatusBadge dentro de uma célula da tabela."""

    def paint(
        self,
        painter: QPainter,
        option: QStyleOptionViewItem,
        index: QModelIndex,
    ) -> None:
        text = index.data(Qt.ItemDataRole.DisplayRole) or ""
        raw = index.data(Qt.ItemDataRole.UserRole)

        _pintar_fundo_sem_texto(self, painter, option, index)

        if not text:
            return

        # Decide cor pelo status
        try:
            status = (
                raw if isinstance(raw, BadgeStatus)
                else BadgeStatus(str(raw).lower())
            )
            cor = _BADGE_CORES[status]
        except (ValueError, KeyError):
            cor = QColor(0x78, 0x7A, 0x80)

        painter.save()
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        f = QFont(option.font)
        f.setPointSize(9)
        f.setWeight(QFont.Weight.Medium)
        f.setCapitalization(QFont.Capitalization.AllUppercase)
        painter.setFont(f)

        fm = painter.fontMetrics()
        text_w = fm.horizontalAdvance(text)
        pill_w = text_w + 16
        pill_h = 20

        cx = option.rect.left() + 8
        cy = option.rect.center().y() - pill_h // 2

        from PySide6.QtCore import QRect
        pill = QRect(cx, cy, pill_w, pill_h)

        fundo = QColor(cor)
        fundo.setAlphaF(0.12)
        painter.setBrush(fundo)
        from PySide6.QtGui import QPen
        painter.setPen(QPen(cor, 1))
        painter.drawRoundedRect(pill, 2, 2)

        painter.setPen(QColor(0x53, 0x56, 0x5A))
        painter.drawText(pill, Qt.AlignmentFlag.AlignCenter, text)
        painter.restore()


# --------------------------------------------------------------------
# Delegate de CodeLink (pinta linha "L{n}" como link teal)
# --------------------------------------------------------------------

class _CodeLinkCellDelegate(QStyledItemDelegate):
    def paint(
        self,
        painter: QPainter,
        option: QStyleOptionViewItem,
        index: QModelIndex,
    ) -> None:
        text = index.data(Qt.ItemDataRole.DisplayRole) or ""
        _pintar_fundo_sem_texto(self, painter, option, index)

        if not text:
            return

        painter.save()
        f = QFont(option.font)
        f.setUnderline(True)
        painter.setFont(f)
        painter.setPen(QColor(0x00, 0x8C, 0x95))
        rect = option.rect.adjusted(8, 0, -8, 0)
        painter.drawText(
            rect,
            Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
            text,
        )
        painter.restore()


# --------------------------------------------------------------------
# DataTable (widget composto)
# --------------------------------------------------------------------

class DataTable(QWidget):
    """Tabela densa com filtro, ordenação, seleção e exportação Excel.

    API pública mínima:
      - construtor: DataTable(columns, rows, *, with_search=True, parent=None)
      - signals: row_activated(dict), row_selected(dict), export_requested(list[dict])
      - métodos: set_rows, add_row, update_row, remove_rows, set_filter,
                 selected_row, visible_rows, set_search_visible
    """

    row_activated = Signal(dict)
    row_selected = Signal(dict)
    export_requested = Signal(list)

    def __init__(
        self,
        columns: list[ColumnSpec],
        rows: list[dict] | None = None,
        *,
        with_search: bool = True,
        with_export: bool = True,
        empty_message: str = "Nenhum registro a exibir.",
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._columns = list(columns)
        self._empty_message = empty_message
        self._with_export = with_export

        v = QVBoxLayout(self)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(8)

        # --- Top bar (search + count) ----------------------------------
        if with_search:
            self._search = SearchField(placeholder=self._search_placeholder(rows or []))
            self._search.query_changed.connect(self.set_filter)
            self._search.cleared.connect(lambda: self.set_filter(""))
            v.addWidget(self._search)
        else:
            self._search = None

        # --- Modelo + proxy + view ----------------------------------
        self._model = _TableModel(self._columns, rows or [], self)
        self._proxy = QSortFilterProxyModel(self)
        self._proxy.setSourceModel(self._model)
        self._proxy.setFilterCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self._proxy.setFilterKeyColumn(-1)  # filtra todas as colunas
        self._proxy.setSortRole(Qt.ItemDataRole.UserRole)

        self._view = QTableView(self)
        self._view.setModel(self._proxy)
        self._view.setSortingEnabled(True)
        self._view.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._view.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._view.setAlternatingRowColors(True)
        self._view.setShowGrid(False)
        self._view.verticalHeader().setVisible(False)
        self._view.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._view.setStyleSheet(self._view_qss())

        hh = self._view.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hh.setStretchLastSection(True)
        hh.setSectionsMovable(True)  # decisão #35: permitir reorder por drag
        hh.setHighlightSections(False)

        self._aplicar_delegates_e_larguras()

        v.addWidget(self._view, 1)

        # --- Empty state label sobreposto ---------------------------
        self._empty_label = QLabel(self._empty_message, self._view)
        self._empty_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._empty_label.setStyleSheet(
            "color: #787A80; font-size: 11pt; background: transparent;"
        )
        self._empty_label.setVisible(False)

        # --- Footer (contador + export) -----------------------------
        footer = QHBoxLayout()
        footer.setContentsMargins(0, 0, 0, 0)

        self._counter = QLabel("")
        self._counter.setStyleSheet("color: #787A80; font-size: 9pt;")
        footer.addWidget(self._counter)
        footer.addStretch()

        if with_export:
            self._export_btn = QPushButton("Exportar Excel")
            self._export_btn.setShortcut(QKeySequence("Ctrl+E"))
            self._export_btn.setStyleSheet(self._export_btn_qss())
            self._export_btn.clicked.connect(self._on_export_clicked)
            footer.addWidget(self._export_btn)

        v.addLayout(footer)

        # --- Conexões ------------------------------------------------
        self._view.activated.connect(self._on_activated)
        sel_model = self._view.selectionModel()
        sel_model.currentRowChanged.connect(self._on_current_row_changed)

        self._proxy.layoutChanged.connect(self._atualizar_contador_e_empty)
        self._model.modelReset.connect(self._atualizar_contador_e_empty)
        self._model.rowsInserted.connect(self._atualizar_contador_e_empty)
        self._model.rowsRemoved.connect(self._atualizar_contador_e_empty)

        # Atalho global Ctrl+E (mesmo sem botão)
        if not with_export:
            QShortcut(QKeySequence("Ctrl+E"), self, activated=self._on_export_clicked)

        self._atualizar_contador_e_empty()

    # ------------------------------------------------------------
    # API pública
    # ------------------------------------------------------------

    def set_rows(self, rows: list[dict]) -> None:
        self._model.set_rows(rows)
        if self._search:
            self._search.set_match_count(len(rows), len(rows))
            self._search._edit.setPlaceholderText(self._search_placeholder(rows))

    def add_row(self, row: dict) -> None:
        self._model.add_row(row)

    def update_row(self, predicate: Callable[[dict], bool], patch: dict) -> int:
        return self._model.update_row(predicate, patch)

    def remove_rows(self, predicate: Callable[[dict], bool]) -> int:
        return self._model.remove_rows(predicate)

    def set_filter(self, query: str) -> None:
        self._proxy.setFilterFixedString(query or "")
        self._atualizar_contador_e_empty()

    def selected_row(self) -> dict | None:
        idx = self._view.currentIndex()
        if not idx.isValid():
            return None
        src = self._proxy.mapToSource(idx)
        if not src.isValid():
            return None
        return self._model.rows()[src.row()]

    def visible_rows(self) -> list[dict]:
        rows = []
        for r in range(self._proxy.rowCount()):
            src = self._proxy.mapToSource(self._proxy.index(r, 0))
            if src.isValid():
                rows.append(self._model.rows()[src.row()])
        return rows

    def total_rows(self) -> int:
        return self._model.rowCount()

    def visible_count(self) -> int:
        return self._proxy.rowCount()

    # ------------------------------------------------------------
    # Eventos / signals
    # ------------------------------------------------------------

    def _on_activated(self, _proxy_index: QModelIndex) -> None:
        row = self.selected_row()
        if row is not None:
            self.row_activated.emit(row)

    def _on_current_row_changed(self, current: QModelIndex, _previous: QModelIndex) -> None:
        if not current.isValid():
            return
        src = self._proxy.mapToSource(current)
        if not src.isValid():
            return
        row = self._model.rows()[src.row()]
        self.row_selected.emit(row)

    def _on_export_clicked(self) -> None:
        """Sempre emite o signal. Consumidor decide se chama dialog default
        via `dt.exportar_via_dialog()` ou se gera arquivo internamente."""
        self.export_requested.emit(self.visible_rows())

    def exportar_via_dialog(self) -> None:
        """Helper público: abre QFileDialog e grava XLSX da view atual.

        Para uso standalone (sem consumidor conectado a export_requested).
        """
        rows = self.visible_rows()
        if not rows:
            return
        self._exportar_default(rows)

    def _exportar_default(self, rows: list[dict]) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar para Excel",
            "export.xlsx",
            "Planilha Excel (*.xlsx)",
        )
        if not path:
            return
        try:
            self._gerar_xlsx(Path(path), rows)
        except Exception as exc:  # noqa: BLE001
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Erro ao exportar", str(exc))

    def _gerar_xlsx(self, destino: Path, rows: list[dict]) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Diagnóstico"

        for c, col in enumerate(self._columns, start=1):
            cel = ws.cell(row=1, column=c, value=col.header)
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="008C95")

        for r, row in enumerate(rows, start=2):
            for c, col in enumerate(self._columns, start=1):
                v = row.get(col.id)
                if col.kind == "money":
                    try:
                        v = float(Decimal(str(v))) if v is not None else 0.0
                    except (InvalidOperation, ValueError, TypeError):
                        v = str(v) if v is not None else ""
                elif col.kind == "code_link" and isinstance(v, dict):
                    v = v.get("linha", "")
                elif isinstance(v, BadgeStatus):
                    v = v.value
                ws.cell(row=r, column=c, value=v)

        wb.save(str(destino))

    # ------------------------------------------------------------
    # Internos — render
    # ------------------------------------------------------------

    def _aplicar_delegates_e_larguras(self) -> None:
        for c, col in enumerate(self._columns):
            if col.kind == "money":
                self._view.setItemDelegateForColumn(
                    c, MoneyCellDelegate(col.money_options or MoneyOptions(), self)
                )
            elif col.kind == "badge":
                self._view.setItemDelegateForColumn(c, _BadgeCellDelegate(self))
            elif col.kind == "code_link":
                self._view.setItemDelegateForColumn(c, _CodeLinkCellDelegate(self))
            if col.width is not None:
                self._view.setColumnWidth(c, col.width)

    def _atualizar_contador_e_empty(self) -> None:
        total = self._model.rowCount()
        visiveis = self._proxy.rowCount()
        if total == 0:
            self._counter.setText("")
            self._empty_label.setText(self._empty_message)
        elif visiveis < total:
            self._counter.setText(f"Mostrando {visiveis} de {total}")
        else:
            self._counter.setText(f"{total} registros")

        self._empty_label.setVisible(visiveis == 0)
        if visiveis == 0:
            # Posiciona o label no centro da viewport
            vp = self._view.viewport()
            self._empty_label.setGeometry(0, 0, vp.width(), vp.height())
            self._empty_label.raise_()

        if self._search is not None and total > 0:
            self._search.set_match_count(visiveis, total)

    def resizeEvent(self, ev) -> None:  # noqa: N802
        super().resizeEvent(ev)
        if self._empty_label.isVisible():
            vp = self._view.viewport()
            self._empty_label.setGeometry(0, 0, vp.width(), vp.height())

    def _search_placeholder(self, rows: list[dict]) -> str:
        return f"Filtrar {len(rows)} {'registro' if len(rows) == 1 else 'registros'}..."

    @staticmethod
    def _view_qss() -> str:
        return """
        QTableView {
            background: #FFFFFF;
            alternate-background-color: #F7F7F8;
            color: #53565A;
            gridline-color: #D1D3D6;
            font-size: 10pt;
            selection-background-color: #E6F3F4;
            selection-color: #53565A;
            border: 1px solid #D1D3D6;
        }
        QTableView::item {
            padding: 4px 8px;
        }
        QTableView::item:selected {
            background: #E6F3F4;
            color: #53565A;
            border-left: 2px solid #008C95;
        }
        QHeaderView::section {
            background: #008C95;
            color: #FFFFFF;
            font-weight: 600;
            padding: 6px 8px;
            border: 0;
            border-right: 1px solid #006F76;
        }
        QHeaderView::section:hover {
            background: #00A4AE;
        }
        """

    @staticmethod
    def _export_btn_qss() -> str:
        return """
        QPushButton {
            background: #FFFFFF;
            color: #008C95;
            border: 1px solid #008C95;
            border-radius: 2px;
            padding: 4px 12px;
            font-size: 9pt;
            font-weight: 500;
        }
        QPushButton:hover {
            background: #E6F3F4;
        }
        QPushButton:pressed {
            background: #B3D7DA;
        }
        """

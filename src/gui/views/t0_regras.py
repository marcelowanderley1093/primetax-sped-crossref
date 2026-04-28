"""
T0 — Visualizador de regras de cruzamento (read-only).

Lista as 49 regras ativas no motor (`src.crossref.engine`) com seus
metadados: código, camada, descrição, base legal, dependências de
SPED e sprint de origem. Permite ao auditor entender o que o sistema
cobre antes de rodar qualquer diagnóstico.

Não permite edição — adicionar regra é fluxo de código (slash command
`/novo-cruzamento` → revisão → merge), conforme CLAUDE.md §11.
"""

from __future__ import annotations

import logging
from pathlib import Path

from PySide6.QtCore import QStandardPaths, Qt, QUrl, Signal
from PySide6.QtGui import QDesktopServices
from PySide6.QtWidgets import (
    QApplication,
    QFrame,
    QHBoxLayout,
    QLabel,
    QProgressDialog,
    QPushButton,
    QScrollArea,
    QTextBrowser,
    QVBoxLayout,
    QWidget,
)

from src.gui.controllers.regras_controller import RegraInfo, RegrasController
from src.gui.widgets import (
    BadgeStatus,
    BreadcrumbSegment,
    ColumnSpec,
    DataTable,
    StatCard,
    Toast,
    ToastAction,
    TraceabilityBreadcrumb,
)


logger = logging.getLogger(__name__)


_CAMADAS_LABEL = {
    1: "Integridade",
    2: "Oportunidade",
    3: "Consistência",
}

_SEVERIDADE_PARA_BADGE = {
    "alto": BadgeStatus.ALTO,
    "medio": BadgeStatus.MEDIO,
    "médio": BadgeStatus.MEDIO,
    "baixo": BadgeStatus.BAIXO,
    "ok": BadgeStatus.OK,
    "n/a": BadgeStatus.NA,
    "—": BadgeStatus.NA,
    "": BadgeStatus.NA,
}


class T0Regras(QWidget):
    """Visualizador read-only das 49 regras de cruzamento."""

    csv_exportado = Signal(Path)

    def __init__(
        self,
        controller: RegrasController | None = None,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._controller = controller or RegrasController()
        self._regras: list[RegraInfo] = []
        self._regra_selecionada: RegraInfo | None = None

        v = QVBoxLayout(self)
        v.setContentsMargins(20, 16, 20, 16)
        v.setSpacing(10)

        # Header com botão de export à direita
        header_row = QHBoxLayout()
        header_row.setSpacing(12)
        self._titulo = QLabel("Regras de Cruzamento")
        self._titulo.setStyleSheet(
            "color: #008C95; font-size: 18pt; font-weight: 600;"
        )
        header_row.addWidget(self._titulo, 1)

        self._btn_export = QPushButton("⬇ Exportar Regras em Excel")
        self._btn_export.setStyleSheet(self._qss_btn_primario())
        self._btn_export.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_export.clicked.connect(self._exportar_regras)
        header_row.addWidget(self._btn_export, 0, Qt.AlignmentFlag.AlignRight)

        wrap_header = QWidget()
        wrap_header.setLayout(header_row)
        v.addWidget(wrap_header)

        self._breadcrumb = TraceabilityBreadcrumb()
        self._breadcrumb.set_segments([
            BreadcrumbSegment(label="Home", target_tela="T1"),
            BreadcrumbSegment(label="Regras"),
        ])
        v.addWidget(self._breadcrumb)

        # Cards de resumo (totais por camada)
        cards_row = QHBoxLayout()
        cards_row.setSpacing(12)
        self._card_total = StatCard(title="Total de regras")
        self._card_camada1 = StatCard(title="Camada 1 (integridade)")
        self._card_camada2 = StatCard(title="Camada 2 (oportunidade)")
        self._card_camada3 = StatCard(title="Camada 3 (consistência)")
        for c in (
            self._card_total, self._card_camada1,
            self._card_camada2, self._card_camada3,
        ):
            cards_row.addWidget(c)
        wrap_cards = QWidget()
        wrap_cards.setLayout(cards_row)
        v.addWidget(wrap_cards)

        # Corpo: tabela à esquerda + painel detalhe à direita
        corpo = QHBoxLayout()
        corpo.setSpacing(12)

        self._tabela = DataTable(
            columns=self._construir_colunas(),
            rows=[],
            with_search=True,
            with_export=True,
            empty_message="Nenhuma regra encontrada (motor vazio?).",
        )
        self._tabela.setMinimumHeight(360)
        self._tabela.row_selected.connect(self._on_regra_selecionada)
        # DataTable emite signal export_requested quando clica no botão
        # interno ou Ctrl+E — conecta no nosso handler que vai pra Downloads.
        self._tabela.export_requested.connect(lambda _: self._exportar_regras())
        corpo.addWidget(self._tabela, 3)

        # Painel detalhe — usa QTextBrowser em vez de QLabel porque o
        # QLabel não quebra strings longas (ex: REGRA_COMPATIBILIDADE_K155_E155
        # ou caminhos Python como src.crossref.camada_2_oportunidades.cruzamento_X)
        # e força overflow horizontal clipado. QTextBrowser quebra
        # naturalmente e tem scroll vertical embutido.
        painel = QWidget()
        painel.setStyleSheet(
            "background: #F7F7F8; border: 1px solid #D1D3D6; border-radius: 4px;"
        )
        pv = QVBoxLayout(painel)
        pv.setContentsMargins(14, 12, 14, 12)
        pv.setSpacing(8)

        self._painel_titulo = QLabel("Detalhe da regra")
        self._painel_titulo.setStyleSheet(
            "color: #53565A; font-size: 12pt; font-weight: 600; "
            "background: transparent; border: none;"
        )
        pv.addWidget(self._painel_titulo)

        self._painel_corpo = QTextBrowser()
        self._painel_corpo.setOpenExternalLinks(False)
        self._painel_corpo.setHtml(
            "<p style='color:#53565A'>"
            "Selecione uma regra na tabela para ver código, base legal, "
            "dependências de SPED e sprint de origem."
            "</p>"
        )
        # Força wrap em qualquer caractere quando não houver espaço — evita
        # overflow horizontal causado por identificadores longos.
        self._painel_corpo.setLineWrapMode(QTextBrowser.LineWrapMode.WidgetWidth)
        self._painel_corpo.setStyleSheet(
            "QTextBrowser { background: #FFFFFF; "
            "border: 1px solid #E5E6E8; border-radius: 2px; "
            "color: #53565A; font-size: 10pt; padding: 8px; }"
        )
        pv.addWidget(self._painel_corpo, 1)

        painel.setMinimumWidth(340)
        corpo.addWidget(painel, 2)

        wrap_corpo = QWidget()
        wrap_corpo.setLayout(corpo)
        v.addWidget(wrap_corpo, 1)

        self._carregar_regras()

    # ------------------------------------------------------------
    # Internos
    # ------------------------------------------------------------

    def _carregar_regras(self) -> None:
        try:
            self._regras = self._controller.listar_regras()
        except Exception as exc:  # noqa: BLE001
            logger.exception("falha ao introspectar regras")
            self._regras = []

        rows = []
        for r in self._regras:
            rows.append({
                "_codigo": r.codigo,
                "codigo": r.codigo,
                "camada": _CAMADAS_LABEL.get(r.camada, str(r.camada)),
                "descricao": r.descricao,
                "severidade": _SEVERIDADE_PARA_BADGE.get(
                    r.severidade.lower(), BadgeStatus.NA
                ),
                "deps": ", ".join(r.dependencias_sped),
                "sprint": r.sprint or "—",
            })
        self._tabela.set_rows(rows)
        self._popular_cards()

    def _popular_cards(self) -> None:
        total = len(self._regras)
        c1 = sum(1 for r in self._regras if r.camada == 1)
        c2 = sum(1 for r in self._regras if r.camada == 2)
        c3 = sum(1 for r in self._regras if r.camada == 3)
        self._card_total.set_primary_value(str(total))
        self._card_total.set_secondary_value("regras ativas no motor", style="normal")
        self._card_camada1.set_primary_value(str(c1))
        self._card_camada2.set_primary_value(str(c2))
        self._card_camada3.set_primary_value(str(c3))

    def _on_regra_selecionada(self, row_dict: dict) -> None:
        codigo = row_dict.get("_codigo")
        regra = next(
            (r for r in self._regras if r.codigo == codigo), None,
        )
        self._regra_selecionada = regra
        if regra is None:
            return

        deps_html = ", ".join(
            f"<code>{self._html_escape(d)}</code>" for d in regra.dependencias_sped
        )
        # Quebra o caminho do módulo em pontos pra permitir wrap em
        # viewport estreito (QTextBrowser respeita word-wrap em pontos
        # de quebra naturais — pontos não são, mas espaços invisíveis sim).
        modulo_quebravel = regra.modulo_path.replace(".", ".​")
        self._painel_corpo.setHtml(
            f"<p><b style='font-size:13pt;color:#008C95'>{regra.codigo}</b></p>"
            f"<p><b>{self._html_escape(regra.descricao)}</b></p>"
            f"<p><span style='color:#787A80'>Camada:</span> "
            f"{regra.camada} — {_CAMADAS_LABEL.get(regra.camada, '')}</p>"
            f"<p><span style='color:#787A80'>Sprint de origem:</span> "
            f"{regra.sprint or 'não documentado'}</p>"
            f"<p><span style='color:#787A80'>Severidade típica:</span> "
            f"{regra.severidade or '—'}</p>"
            f"<p><span style='color:#787A80'>SPEDs requeridos:</span> "
            f"{deps_html}</p>"
            f"<p><span style='color:#787A80'>Base legal:</span><br>"
            f"{self._html_escape(regra.base_legal)}</p>"
            f"<p><span style='color:#787A80; font-size:9pt'>Módulo: "
            f"<code>{self._html_escape(modulo_quebravel)}</code></span></p>"
        )

    # ------------------------------------------------------------
    # Export Excel — pasta Downloads do usuário
    # ------------------------------------------------------------

    def _exportar_regras(self) -> None:
        """Exporta as 49 regras pra ~/Downloads/primetax_regras_de_cruzamento.xlsx
        com cabeçalho institucional e todas as colunas (código, camada,
        descrição, severidade, SPEDs, sprint, base legal, módulo)."""
        if not self._regras:
            Toast.show_warning(
                self.window(), "Sem regras para exportar.",
            )
            return
        downloads = QStandardPaths.writableLocation(
            QStandardPaths.StandardLocation.DownloadLocation
        )
        base = Path(downloads) if downloads else Path.home() / "Downloads"
        base.mkdir(parents=True, exist_ok=True)
        destino = base / "primetax_regras_de_cruzamento.xlsx"

        dialog = QProgressDialog(
            "Exportando Regras de Cruzamento...\n\n"
            "Aguarde alguns instantes.",
            "", 0, 0, self.window(),
        )
        dialog.setWindowTitle("Primetax SPED — Exportando")
        dialog.setCancelButton(None)
        dialog.setWindowModality(Qt.WindowModality.ApplicationModal)
        dialog.setMinimumDuration(0)
        dialog.setMinimumWidth(360)
        dialog.show()
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        QApplication.processEvents()
        QApplication.processEvents()

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = Workbook()
            ws = wb.active
            ws.title = "Regras de Cruzamento"

            ws.cell(
                row=1, column=1,
                value="REGRAS DE CRUZAMENTO — PRIMETAX SPED",
            ).font = Font(bold=True, size=14, color="008C95")
            ws.cell(
                row=2, column=1,
                value=f"Total: {len(self._regras)} regras ativas no motor",
            ).font = Font(italic=True, size=10)

            headers = [
                "Código", "Camada", "Descrição", "Severidade típica",
                "SPEDs requeridos", "Sprint", "Base legal", "Módulo Python",
            ]
            for c, h in enumerate(headers, start=1):
                cel = ws.cell(row=4, column=c, value=h)
                cel.font = Font(bold=True, color="FFFFFF")
                cel.fill = PatternFill("solid", fgColor="008C95")
                cel.alignment = Alignment(horizontal="center", wrap_text=True)

            for r, regra in enumerate(self._regras, start=5):
                ws.cell(row=r, column=1, value=regra.codigo)
                ws.cell(
                    row=r, column=2,
                    value=f"{regra.camada} — {_CAMADAS_LABEL.get(regra.camada, '')}",
                )
                ws.cell(row=r, column=3, value=regra.descricao)
                ws.cell(row=r, column=4, value=regra.severidade or "—")
                ws.cell(row=r, column=5, value=", ".join(regra.dependencias_sped))
                ws.cell(row=r, column=6, value=regra.sprint or "—")
                ws.cell(row=r, column=7, value=regra.base_legal)
                ws.cell(row=r, column=8, value=regra.modulo_path)

            # Larguras de coluna razoáveis
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 24
            ws.column_dimensions["C"].width = 60
            ws.column_dimensions["D"].width = 16
            ws.column_dimensions["E"].width = 30
            ws.column_dimensions["F"].width = 14
            ws.column_dimensions["G"].width = 80
            ws.column_dimensions["H"].width = 60

            wb.save(str(destino))
        except Exception as exc:  # noqa: BLE001
            logger.exception("falha ao exportar regras")
            dialog.close()
            QApplication.restoreOverrideCursor()
            Toast.show_error(
                self.window(), f"Falha ao exportar regras: {exc}",
            )
            return
        finally:
            dialog.close()
            QApplication.restoreOverrideCursor()

        Toast.show_success(
            self.window(),
            f"Regras de cruzamento salvas em: {destino.parent}\\{destino.name}",
            duration_ms=8000,
            action=ToastAction(
                label="Abrir pasta",
                callback=lambda p=destino.parent: QDesktopServices.openUrl(
                    QUrl.fromLocalFile(str(p))
                ),
            ),
        )
        self.csv_exportado.emit(destino)

    # ------------------------------------------------------------

    @staticmethod
    def _qss_btn_primario() -> str:
        return """
        QPushButton {
            background: #008C95; color: #FFFFFF; border: none;
            border-radius: 2px; padding: 6px 14px;
            font-size: 10pt; font-weight: 500;
        }
        QPushButton:hover { background: #00A4AE; }
        QPushButton:pressed { background: #006F76; }
        """

    @staticmethod
    def _construir_colunas() -> list[ColumnSpec]:
        # Descrição é a ÚLTIMA coluna — stretchLastSection=True do DataTable
        # faz ela tomar todo espaço restante, evitando scroll horizontal
        # e garantindo que as outras colunas (códgio, severidade, etc.)
        # fiquem sempre visíveis.
        return [
            ColumnSpec(id="codigo", header="Código", kind="text", width=70),
            ColumnSpec(id="camada", header="Camada", kind="text", width=120),
            ColumnSpec(id="severidade", header="Severidade", kind="badge", width=110),
            ColumnSpec(id="sprint", header="Sprint", kind="text", width=80),
            ColumnSpec(id="deps", header="SPEDs", kind="text", width=160),
            ColumnSpec(id="descricao", header="Descrição", kind="text", width=400),
        ]

    @staticmethod
    def _html_escape(s: str) -> str:
        return (
            (s or "")
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )

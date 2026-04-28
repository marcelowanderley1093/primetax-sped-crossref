"""
T9 — Análise Contábil (BP, DRE, Razão da conta, Despesas × Crédito).

Estrutura tipo "Power BI / livro contábil":
  - BP e DRE em QTreeView hierárquico, colapsado nas sintéticas;
    duplo-clique numa folha (analítica) abre o Razão dessa conta no
    painel horizontal abaixo (QSplitter vertical).
  - Razão tem cabeçalho com saldo inicial, tabela com débito/crédito/
    saldo corrente, e footer com totais e saldo final.
  - Aba "Despesas × Crédito PIS/COFINS" — diferencial pro Tema 779
    (REsp 1.221.170/PR essencialidade). Selecionar uma linha mostra
    no painel embaixo as evidências (F100/F120/F130) com arquivo,
    linha física e registro — rastreabilidade absoluta (CLAUDE.md §1).
  - Cada aba tem seu próprio botão "Exportar Excel" (individualizado).

Pré-requisitos: ECD com J100, J150, I250, I050, I155 importados.
"""

from __future__ import annotations

import logging
from decimal import Decimal
from pathlib import Path

from PySide6.QtCore import QStandardPaths, Qt, QUrl, Signal
from PySide6.QtGui import QDesktopServices, QStandardItem, QStandardItemModel
from PySide6.QtWidgets import (
    QApplication,
    QComboBox,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QProgressDialog,
    QPushButton,
    QSplitter,
    QStackedWidget,
    QTabWidget,
    QTreeView,
    QVBoxLayout,
    QWidget,
)

from src.gui.controllers.clientes_controller import ClienteRow
from src.gui.controllers.contabil_controller import (
    ContabilController,
    DespesaVsCredito,
    EvidenciaCredito,
    ImobilizadoVsCredito,
    LancamentoRazao,
    LinhaBalanco,
    RazaoConta,
)
from src.gui.widgets import (
    BadgeStatus,
    BreadcrumbSegment,
    ColumnSpec,
    DataTable,
    EmptyState,
    InlineMessage,
    MessageLevel,
    Toast,
    ToastAction,
    TraceabilityBreadcrumb,
)


logger = logging.getLogger(__name__)


_PRIMARY_COLOR = "#008C95"


# --------------------------------------------------------------------
# Helpers de hierarquia para QTreeView
# --------------------------------------------------------------------

def _construir_arvore(linhas: list[LinhaBalanco]) -> list[tuple[LinhaBalanco, list]]:
    """Transforma lista plana em estrutura de árvore via cod_agl_sup.
    Retorna [(linha, [filhos_recursivos])...] preservando ordem original."""
    por_pai: dict[str, list[LinhaBalanco]] = {}
    for l in linhas:
        por_pai.setdefault(l.cod_agl_sup or "", []).append(l)

    def construir(pai: str) -> list[tuple[LinhaBalanco, list]]:
        return [
            (l, construir(l.cod_agl))
            for l in por_pai.get(pai, [])
        ]

    return construir("")


def _fmt_brl(v: Decimal) -> str:
    try:
        f = float(v)
    except (TypeError, ValueError):
        return "0,00"
    return f"{f:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_data_brl(s: str) -> str:
    """ISO YYYY-MM-DD → DD/MM/AAAA. Vazio retorna vazio."""
    if not s or len(s) < 10:
        return s or ""
    try:
        ano, mes, dia = s[:4], s[5:7], s[8:10]
        if ano.isdigit() and mes.isdigit() and dia.isdigit():
            return f"{dia}/{mes}/{ano}"
    except Exception:
        pass
    return s


def _fmt_periodo_brl(ano_mes: int) -> str:
    """YYYYMM → MM/AAAA."""
    if not ano_mes:
        return ""
    try:
        ano = ano_mes // 100
        mes = ano_mes % 100
        return f"{mes:02d}/{ano:04d}"
    except Exception:
        return str(ano_mes)


# --------------------------------------------------------------------
# View principal
# --------------------------------------------------------------------

class T9Contabil(QWidget):
    """Painel de análise contábil para um cliente × ano-calendário."""

    csv_exportado = Signal(Path)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._cliente: ClienteRow | None = None
        self._controller: ContabilController | None = None
        self._linhas_bp: list[LinhaBalanco] = []
        self._linhas_dre: list[LinhaBalanco] = []
        self._linhas_despesas: list[DespesaVsCredito] = []
        self._linhas_imobilizado: list[ImobilizadoVsCredito] = []

        v = QVBoxLayout(self)
        v.setContentsMargins(20, 16, 20, 16)
        v.setSpacing(10)

        # Header (sem botão export geral — agora cada aba tem o seu)
        self._titulo = QLabel("Análise Contábil")
        self._titulo.setStyleSheet(
            f"color: {_PRIMARY_COLOR}; font-size: 18pt; font-weight: 600;"
        )
        v.addWidget(self._titulo)

        self._breadcrumb = TraceabilityBreadcrumb()
        v.addWidget(self._breadcrumb)

        self._stack = QStackedWidget()
        self._empty = EmptyState(
            title="Nenhum cliente selecionado",
            description=(
                "Selecione um cliente em T1 (Clientes) com ECD importada "
                "para visualizar BP, DRE e razão das contas."
            ),
        )
        self._stack.addWidget(self._empty)

        self._conteudo = QWidget()
        cv = QVBoxLayout(self._conteudo)
        cv.setContentsMargins(0, 0, 0, 0)
        cv.setSpacing(8)

        self._inline_aviso: InlineMessage | None = None
        self._inline_wrapper = QWidget()
        self._inline_layout = QVBoxLayout(self._inline_wrapper)
        self._inline_layout.setContentsMargins(0, 0, 0, 0)
        self._inline_layout.setSpacing(0)
        cv.addWidget(self._inline_wrapper)

        self._tabs = QTabWidget()
        self._tabs.setStyleSheet(self._qss_tabs())
        cv.addWidget(self._tabs, 1)

        self._construir_aba_bp()
        self._construir_aba_dre()
        self._construir_aba_despesas()
        self._construir_aba_imobilizado()

        self._stack.addWidget(self._conteudo)
        v.addWidget(self._stack, 1)

    # ------------------------------------------------------------
    # Construção das abas
    # ------------------------------------------------------------

    def _construir_aba_bp(self) -> None:
        """Aba BP: QTreeView no topo + Razão drill-down embaixo (splitter)."""
        wrapper = QWidget()
        v = QVBoxLayout(wrapper)
        v.setContentsMargins(0, 8, 0, 0)
        v.setSpacing(8)

        # Botão de export individual da aba
        v.addWidget(self._build_export_row("Exportar BP em Excel", self._exportar_bp))

        splitter = QSplitter(Qt.Orientation.Vertical)

        # Topo: árvore do BP
        self._tree_bp = self._build_tree_view()
        self._tree_bp.doubleClicked.connect(self._on_bp_dre_double_click)
        splitter.addWidget(self._tree_bp)

        # Embaixo: razão drill-down (inicialmente vazio)
        self._razao_panel_bp = self._build_razao_panel()
        splitter.addWidget(self._razao_panel_bp)

        splitter.setSizes([500, 0])  # razão fechado por padrão
        splitter.setCollapsible(0, False)
        v.addWidget(splitter, 1)

        self._splitter_bp = splitter
        self._tabs.addTab(wrapper, "Balanço Patrimonial")

    def _construir_aba_dre(self) -> None:
        wrapper = QWidget()
        v = QVBoxLayout(wrapper)
        v.setContentsMargins(0, 8, 0, 0)
        v.setSpacing(8)

        v.addWidget(self._build_export_row("Exportar DRE em Excel", self._exportar_dre))

        splitter = QSplitter(Qt.Orientation.Vertical)

        self._tree_dre = self._build_tree_view()
        self._tree_dre.doubleClicked.connect(self._on_bp_dre_double_click)
        splitter.addWidget(self._tree_dre)

        self._razao_panel_dre = self._build_razao_panel()
        splitter.addWidget(self._razao_panel_dre)

        splitter.setSizes([500, 0])
        splitter.setCollapsible(0, False)
        v.addWidget(splitter, 1)

        self._splitter_dre = splitter
        self._tabs.addTab(wrapper, "DRE")

    def _construir_aba_despesas(self) -> None:
        """Aba Despesas × Crédito com painel inferior em sub-abas:
          - Evidências do Crédito (F100/F120/F130)
          - Razão da Conta (lançamentos do I250 — pra reanalisar Tema 779)
        Ambas populam quando uma despesa é selecionada na tabela superior.
        Auditor pode marcar despesas como oportunidades (Tema 779).
        """
        wrapper = QWidget()
        v = QVBoxLayout(wrapper)
        v.setContentsMargins(0, 8, 0, 0)
        v.setSpacing(8)

        # Linha de export: dois botões (despesas tudo + relatório só marcadas)
        export_row = QHBoxLayout()
        export_row.setContentsMargins(0, 0, 0, 0)
        export_row.addStretch()
        btn_rel = QPushButton("Exportar Relatório de Oportunidades (marcadas)")
        btn_rel.setStyleSheet(self._qss_btn_secundario())
        btn_rel.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_rel.clicked.connect(self._exportar_relatorio_oportunidades)
        export_row.addWidget(btn_rel)
        btn_full = QPushButton("Exportar Despesas × Crédito em Excel")
        btn_full.setStyleSheet(self._qss_btn_secundario())
        btn_full.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_full.clicked.connect(self._exportar_despesas)
        export_row.addWidget(btn_full)
        wrap_export = QWidget()
        wrap_export.setLayout(export_row)
        v.addWidget(wrap_export)

        # Contador de oportunidades marcadas
        self._lbl_contador_oport = QLabel("")
        self._lbl_contador_oport.setStyleSheet(
            "color: #008C95; font-size: 10pt; font-weight: 600;"
            "background: #E6F3F4; border: 1px solid #B3D7DA;"
            "border-radius: 2px; padding: 6px 10px;"
        )
        self._lbl_contador_oport.setVisible(False)
        v.addWidget(self._lbl_contador_oport)

        info = QLabel(
            "Lista contas de despesa do plano (COD_NAT='04') × créditos "
            "PIS/COFINS tomados em F100/F120/F130 da EFD-Contribuições. "
            "Contas com saldo significativo e Status ALTO (zero crédito) "
            "são candidatas à reanálise sob Tema 779 (REsp 1.221.170/PR — "
            "essencialidade). <b>Selecione uma linha</b> para ver as evidências "
            "do crédito e o razão contábil da conta nas abas inferiores."
        )
        info.setWordWrap(True)
        info.setStyleSheet("color: #787A80; font-size: 9pt; padding: 4px 0px;")
        v.addWidget(info)

        splitter = QSplitter(Qt.Orientation.Vertical)

        # Tabela de despesas
        self._tabela_despesas = DataTable(
            columns=self._cols_despesas(),
            rows=[],
            with_search=True,
            with_export=False,
            empty_message="Sem ECD ou sem contas de despesa para este cliente × AC.",
        )
        self._tabela_despesas.setMinimumHeight(220)
        self._tabela_despesas.row_selected.connect(self._on_despesa_selecionada)
        splitter.addWidget(self._tabela_despesas)

        # Painel inferior: sub-abas (evidências + razão)
        sub_tabs = QTabWidget()
        sub_tabs.setStyleSheet(self._qss_tabs())

        # ---- Sub-aba 1: Evidências ----
        evidencias_widget = QWidget()
        ev_v = QVBoxLayout(evidencias_widget)
        ev_v.setContentsMargins(8, 8, 8, 8)
        ev_v.setSpacing(6)

        self._evidencias_header = QLabel(
            "Selecione uma conta acima para ver as evidências do crédito "
            "PIS/COFINS (F100/F120/F130)."
        )
        self._evidencias_header.setWordWrap(True)
        self._evidencias_header.setStyleSheet(
            "color: #53565A; font-size: 11pt; font-weight: 600;"
            "background: #F0F7F8; border: 1px solid #B3D7DA;"
            "border-radius: 2px; padding: 8px 12px;"
        )
        ev_v.addWidget(self._evidencias_header)

        # Linha de ações sobre a conta selecionada (marcar oportunidade)
        acoes_row = QHBoxLayout()
        acoes_row.setContentsMargins(0, 0, 0, 0)
        self._btn_marcar_oport = QPushButton("⚑ Marcar como oportunidade Tema 779")
        self._btn_marcar_oport.setStyleSheet(self._qss_btn_primario())
        self._btn_marcar_oport.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_marcar_oport.clicked.connect(self._marcar_oportunidade)
        self._btn_marcar_oport.setEnabled(False)
        acoes_row.addWidget(self._btn_marcar_oport)
        self._btn_desmarcar_oport = QPushButton("Desmarcar oportunidade")
        self._btn_desmarcar_oport.setStyleSheet(self._qss_btn_secundario())
        self._btn_desmarcar_oport.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_desmarcar_oport.clicked.connect(self._desmarcar_oportunidade)
        self._btn_desmarcar_oport.setEnabled(False)
        acoes_row.addWidget(self._btn_desmarcar_oport)
        acoes_row.addStretch()
        wrap_acoes = QWidget()
        wrap_acoes.setLayout(acoes_row)
        ev_v.addWidget(wrap_acoes)

        self._tabela_evidencias = DataTable(
            columns=self._cols_evidencias(),
            rows=[],
            with_search=False,
            with_export=False,
            empty_message="Nenhuma evidência encontrada para esta conta.",
        )
        self._tabela_evidencias.setMinimumHeight(180)
        ev_v.addWidget(self._tabela_evidencias, 1)

        sub_tabs.addTab(evidencias_widget, "Evidências do Crédito")

        # ---- Sub-aba 2: Razão da conta selecionada ----
        # Reusa o mesmo padrão de painel razão das abas BP/DRE.
        self._razao_panel_despesas = self._build_razao_panel()
        # Wrap em widget para ter padding consistente
        razao_wrap = QWidget()
        rw_v = QVBoxLayout(razao_wrap)
        rw_v.setContentsMargins(8, 8, 8, 8)
        rw_v.setSpacing(0)
        rw_v.addWidget(self._razao_panel_despesas)
        sub_tabs.addTab(razao_wrap, "Razão da Conta")

        splitter.addWidget(sub_tabs)
        splitter.setSizes([320, 320])
        splitter.setCollapsible(0, False)
        v.addWidget(splitter, 1)

        self._tabs.addTab(wrapper, "Despesas × Crédito PIS/COFINS")

    def _construir_aba_imobilizado(self) -> None:
        """Aba Imobilizado × Crédito (complementar à CR-23). Lista contas
        de imobilizado da ECD e cruza com F120 (depreciação) e F130
        (aquisição) da EFD-Contribuições. Auditor marca como oportunidade
        de Tema 779 (essencialidade) e exporta relatório."""
        wrapper = QWidget()
        v = QVBoxLayout(wrapper)
        v.setContentsMargins(0, 8, 0, 0)
        v.setSpacing(8)

        # Linha de export
        export_row = QHBoxLayout()
        export_row.setContentsMargins(0, 0, 0, 0)
        export_row.addStretch()
        btn_rel_imob = QPushButton(
            "Exportar Relatório de Imobilizado (marcadas)"
        )
        btn_rel_imob.setStyleSheet(self._qss_btn_secundario())
        btn_rel_imob.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_rel_imob.clicked.connect(self._exportar_relatorio_imobilizado)
        export_row.addWidget(btn_rel_imob)
        btn_full = QPushButton("Exportar Imobilizado × Crédito em Excel")
        btn_full.setStyleSheet(self._qss_btn_secundario())
        btn_full.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_full.clicked.connect(self._exportar_imobilizado)
        export_row.addWidget(btn_full)
        wrap_export = QWidget()
        wrap_export.setLayout(export_row)
        v.addWidget(wrap_export)

        info = QLabel(
            "Lista contas de <b>imobilizado</b> do plano (cod_nat='01') "
            "com saldo no AC × créditos PIS/COFINS tomados em F120 "
            "(encargos de depreciação) e F130 (aquisição). "
            "Contas com saldo significativo e Status <b>ALTO</b> "
            "(zero crédito tomado) são candidatas. <b>Complementar à CR-23</b>, "
            "que detecta CST 98/99 em C170 — esta análise é mais ampla. "
            "Base legal: art. 3º, VI da Lei 10.637/2002 e 10.833/2003. "
            "<b>Selecione uma linha</b> para ver evidências e razão da conta."
        )
        info.setWordWrap(True)
        info.setStyleSheet("color: #787A80; font-size: 9pt; padding: 4px 0px;")
        v.addWidget(info)

        # Contador de oportunidades imobilizado marcadas
        self._lbl_contador_imob = QLabel("")
        self._lbl_contador_imob.setStyleSheet(
            "color: #008C95; font-size: 10pt; font-weight: 600;"
            "background: #E6F3F4; border: 1px solid #B3D7DA;"
            "border-radius: 2px; padding: 6px 10px;"
        )
        self._lbl_contador_imob.setVisible(False)
        v.addWidget(self._lbl_contador_imob)

        splitter = QSplitter(Qt.Orientation.Vertical)

        # Tabela de imobilizado
        self._tabela_imobilizado = DataTable(
            columns=self._cols_imobilizado(),
            rows=[],
            with_search=True,
            with_export=False,
            empty_message="Sem ECD ou sem contas de imobilizado para este cliente × AC.",
        )
        self._tabela_imobilizado.setMinimumHeight(220)
        self._tabela_imobilizado.row_selected.connect(self._on_imobilizado_selecionado)
        splitter.addWidget(self._tabela_imobilizado)

        # Painel inferior: ações + razão
        painel_inf = QWidget()
        pi_v = QVBoxLayout(painel_inf)
        pi_v.setContentsMargins(8, 8, 8, 8)
        pi_v.setSpacing(6)

        self._imob_header = QLabel(
            "Selecione uma conta acima para ver o razão e ações."
        )
        self._imob_header.setWordWrap(True)
        self._imob_header.setStyleSheet(
            "color: #53565A; font-size: 11pt; font-weight: 600;"
            "background: #F0F7F8; border: 1px solid #B3D7DA;"
            "border-radius: 2px; padding: 8px 12px;"
        )
        pi_v.addWidget(self._imob_header)

        # Botões de marcação (mesmos métodos de despesa, mesma tabela DB)
        acoes_row = QHBoxLayout()
        acoes_row.setContentsMargins(0, 0, 0, 0)
        self._btn_marcar_imob = QPushButton("⚑ Marcar como oportunidade Tema 779")
        self._btn_marcar_imob.setStyleSheet(self._qss_btn_primario())
        self._btn_marcar_imob.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_marcar_imob.clicked.connect(self._marcar_oportunidade_imob)
        self._btn_marcar_imob.setEnabled(False)
        acoes_row.addWidget(self._btn_marcar_imob)
        self._btn_desmarcar_imob = QPushButton("Desmarcar oportunidade")
        self._btn_desmarcar_imob.setStyleSheet(self._qss_btn_secundario())
        self._btn_desmarcar_imob.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_desmarcar_imob.clicked.connect(self._desmarcar_oportunidade_imob)
        self._btn_desmarcar_imob.setEnabled(False)
        acoes_row.addWidget(self._btn_desmarcar_imob)
        acoes_row.addStretch()
        wrap_acoes = QWidget()
        wrap_acoes.setLayout(acoes_row)
        pi_v.addWidget(wrap_acoes)

        # Razão drill-down (mesmo padrão das outras abas)
        self._razao_panel_imob = self._build_razao_panel()
        pi_v.addWidget(self._razao_panel_imob, 1)

        splitter.addWidget(painel_inf)
        splitter.setSizes([280, 360])
        splitter.setCollapsible(0, False)
        v.addWidget(splitter, 1)

        self._tabs.addTab(wrapper, "Imobilizado × Crédito PIS/COFINS")

    def _build_export_row(self, label: str, callback) -> QWidget:
        row = QHBoxLayout()
        row.setContentsMargins(0, 0, 0, 0)
        row.addStretch()
        btn = QPushButton(label)
        btn.setStyleSheet(self._qss_btn_secundario())
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        btn.clicked.connect(callback)
        row.addWidget(btn)
        wrap = QWidget()
        wrap.setLayout(row)
        return wrap

    def _build_tree_view(self) -> QTreeView:
        tv = QTreeView()
        tv.setHeaderHidden(False)
        tv.setAlternatingRowColors(True)
        tv.setRootIsDecorated(True)
        tv.setUniformRowHeights(True)
        tv.setStyleSheet(self._qss_tree())
        return tv

    def _build_razao_panel(self) -> QWidget:
        """Painel de razão drill-down: header com botão de export +
        tabela + footer. Botão fica no header (lado direito) pra estar
        sempre visível, em vez de no rodapé que pode sumir em splitter."""
        panel = QWidget()
        v = QVBoxLayout(panel)
        v.setContentsMargins(0, 4, 0, 0)
        v.setSpacing(6)

        # Linha de header: label + botão Exportar Razão (lado direito)
        header_row = QHBoxLayout()
        header_row.setContentsMargins(0, 0, 0, 0)
        header_row.setSpacing(8)

        header = QLabel("")
        header.setWordWrap(True)
        header.setStyleSheet(
            "color: #53565A; font-size: 11pt; font-weight: 600;"
            "background: #F0F7F8; border: 1px solid #B3D7DA;"
            "border-radius: 2px; padding: 8px 12px;"
        )
        header.setVisible(False)
        header_row.addWidget(header, 1)

        btn_exp = QPushButton("⬇ Exportar Razão em Excel")
        btn_exp.setStyleSheet(self._qss_btn_primario())
        btn_exp.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_exp.clicked.connect(self._exportar_razao_atual)
        btn_exp.setEnabled(False); btn_exp.setVisible(True)
        btn_exp.setVisible(False)  # só aparece quando há dados de razão
        header_row.addWidget(btn_exp, 0, Qt.AlignmentFlag.AlignTop)

        wrap_header = QWidget()
        wrap_header.setLayout(header_row)
        v.addWidget(wrap_header)

        tabela = DataTable(
            columns=self._cols_razao(),
            rows=[],
            with_search=True,
            with_export=False,
            empty_message="Duplo-clique numa conta analítica acima.",
        )
        tabela.setMinimumHeight(180)
        v.addWidget(tabela, 1)

        footer = QLabel("")
        footer.setWordWrap(True)
        footer.setStyleSheet(
            "color: #53565A; font-size: 11pt; font-weight: 600;"
            "background: #F7F7F8; border: 1px solid #D1D3D6;"
            "border-top: 2px solid #008C95;"
            "padding: 8px 12px;"
        )
        footer.setVisible(False)
        v.addWidget(footer)

        # Guarda referências por convenção (usadas em _abrir_razao)
        panel._header = header  # type: ignore[attr-defined]
        panel._tabela = tabela  # type: ignore[attr-defined]
        panel._footer = footer  # type: ignore[attr-defined]
        panel._btn_export = btn_exp  # type: ignore[attr-defined]
        panel._cod_cta_atual = ""  # type: ignore[attr-defined]

        return panel

    # ------------------------------------------------------------
    # API pública
    # ------------------------------------------------------------

    def carregar_cliente(self, cliente: ClienteRow) -> None:
        self._cliente = cliente
        self._controller = ContabilController(
            cnpj=cliente.cnpj,
            ano_calendario=cliente.ano_calendario,
        )
        self._titulo.setText(
            f"Análise Contábil · {cliente.razao_social} · AC {cliente.ano_calendario}"
        )
        self._breadcrumb.set_segments([
            BreadcrumbSegment(label="Home", target_tela="T1"),
            BreadcrumbSegment(
                label=f"{cliente.razao_social} × {cliente.ano_calendario}",
                target_tela="T3",
            ),
            BreadcrumbSegment(label="Análise Contábil"),
        ])
        self._stack.setCurrentWidget(self._conteudo)
        self._recarregar()

    def cliente_atual(self) -> ClienteRow | None:
        return self._cliente

    # ------------------------------------------------------------
    # Carregamento
    # ------------------------------------------------------------

    def _recarregar(self) -> None:
        if self._controller is None:
            return
        disp = self._controller.disponibilidade()
        self._mostrar_aviso(disp)

        # BP
        if disp.tem_j100:
            self._linhas_bp = self._controller.listar_balanco_patrimonial()
            self._popular_tree(self._tree_bp, self._linhas_bp, modo="bp")
        else:
            self._linhas_bp = []
            self._tree_bp.setModel(None)

        # DRE
        if disp.tem_j150:
            self._linhas_dre = self._controller.listar_dre()
            self._popular_tree(self._tree_dre, self._linhas_dre, modo="dre")
        else:
            self._linhas_dre = []
            self._tree_dre.setModel(None)

        # Despesas
        if disp.tem_ecd:
            self._linhas_despesas = self._controller.listar_despesas_vs_credito()
            self._tabela_despesas.set_rows(
                [self._despesa_para_dict(d) for d in self._linhas_despesas]
            )
            self._atualizar_contador_oportunidades()
            # Imobilizado também depende de ECD + (idealmente) EFD-Contrib
            self._linhas_imobilizado = self._controller.listar_imobilizado_vs_credito()
            self._tabela_imobilizado.set_rows(
                [self._imobilizado_para_dict(d) for d in self._linhas_imobilizado]
            )
            self._atualizar_contador_imobilizado()
        else:
            self._linhas_despesas = []
            self._tabela_despesas.set_rows([])
            self._lbl_contador_oport.setVisible(False)
            self._linhas_imobilizado = []
            self._tabela_imobilizado.set_rows([])
            self._lbl_contador_imob.setVisible(False)

        # Razão fechado por padrão
        self._splitter_bp.setSizes([500, 0])
        self._splitter_dre.setSizes([500, 0])

    def _popular_tree(
        self, tree: QTreeView, linhas: list[LinhaBalanco], *, modo: str,
    ) -> None:
        """Preenche QTreeView com hierarquia. Linhas sintéticas em negrito.
        Valores formatados com R$. Cliente clica para expandir; duplo-clique
        em folha (analítica) abre razão."""
        from PySide6.QtGui import QBrush, QColor, QFont

        model = QStandardItemModel()
        model.setHorizontalHeaderLabels(["Descrição", "Saldo Final"])

        arvore = _construir_arvore(linhas)
        for linha, filhos in arvore:
            item_descr, item_valor = self._tree_items(linha, modo=modo)
            self._adicionar_filhos(item_descr, item_valor, filhos, modo=modo)
            model.appendRow([item_descr, item_valor])

        tree.setModel(model)
        # Coluna Descrição grande, Valor menor à direita
        tree.header().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        tree.header().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        tree.header().setStretchLastSection(False)
        # Expande primeiro nível só
        for r in range(model.rowCount()):
            tree.setExpanded(model.index(r, 0), True)

    def _adicionar_filhos(
        self, parent_descr: QStandardItem, parent_valor: QStandardItem,
        filhos: list[tuple[LinhaBalanco, list]], *, modo: str,
    ) -> None:
        for linha, netos in filhos:
            item_descr, item_valor = self._tree_items(linha, modo=modo)
            self._adicionar_filhos(item_descr, item_valor, netos, modo=modo)
            parent_descr.appendRow([item_descr, item_valor])

    def _tree_items(
        self, linha: LinhaBalanco, *, modo: str,
    ) -> tuple[QStandardItem, QStandardItem]:
        from PySide6.QtGui import QFont
        sintetica = (linha.ind_cod_agl or "").upper() == "T"

        # Descrição
        item_d = QStandardItem(linha.descricao)
        item_d.setEditable(False)
        # Guarda metadados pra drill-down
        item_d.setData(linha.cod_agl, Qt.ItemDataRole.UserRole)
        item_d.setData(linha.descricao, Qt.ItemDataRole.UserRole + 1)
        item_d.setData(sintetica, Qt.ItemDataRole.UserRole + 2)

        # Valor (alinhado à direita)
        valor = self._valor_signed(linha, modo=modo)
        valor_abs = abs(valor)
        prefixo = "(" if valor < 0 else ""
        sufixo = ")" if valor < 0 else ""
        texto_valor = f"{prefixo}R$ {_fmt_brl(valor_abs)}{sufixo}"
        item_v = QStandardItem(texto_valor)
        item_v.setEditable(False)
        item_v.setTextAlignment(
            Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
        )

        # Estilo: sintéticas em negrito, valores negativos em vermelho discreto
        if sintetica:
            f = QFont()
            f.setBold(True)
            item_d.setFont(f)
            item_v.setFont(f)

        return item_d, item_v

    @staticmethod
    def _valor_signed(linha: LinhaBalanco, *, modo: str) -> Decimal:
        """Valor com sinal contábil para apresentação:
          - BP Ativo (A): D=positivo, C=negativo
          - BP Passivo+PL (P): C=positivo, D=negativo
          - DRE: C=positivo (receitas), D=negativo (custos/despesas)
        """
        if modo == "bp":
            if linha.grupo == "A":
                return linha.vl_final if linha.ind_dc_fin == "D" else -linha.vl_final
            return linha.vl_final if linha.ind_dc_fin == "C" else -linha.vl_final
        # DRE
        return linha.vl_final if linha.ind_dc_fin == "C" else -linha.vl_final

    def _mostrar_aviso(self, disp) -> None:
        if self._inline_aviso is not None:
            self._inline_aviso.deleteLater()
            self._inline_aviso = None

        if not disp.tem_ecd:
            self._inline_aviso = InlineMessage(
                MessageLevel.WARNING,
                "Este cliente × AC não tem ECD importada. A análise contábil "
                "(BP, DRE, razão) depende da ECD — importe-a em T2.",
            )
            self._inline_layout.addWidget(self._inline_aviso)

    # ------------------------------------------------------------
    # Drill-down: BP/DRE folha → Razão
    # ------------------------------------------------------------

    def _on_bp_dre_double_click(self, idx) -> None:
        """Duplo-clique em conta analítica (folha) abre razão no painel
        horizontal abaixo da árvore. Sintéticas só expandem/colapsam."""
        if not idx.isValid():
            return
        col0 = idx.sibling(idx.row(), 0)
        sintetica = col0.data(Qt.ItemDataRole.UserRole + 2)
        cod_agl = col0.data(Qt.ItemDataRole.UserRole)
        descr = col0.data(Qt.ItemDataRole.UserRole + 1)
        if sintetica:
            # Sintética: deixa o Qt expandir/colapsar default
            return
        # Determina qual splitter (BP ou DRE) acionar
        sender = self.sender()
        if sender is self._tree_bp:
            self._abrir_razao(self._razao_panel_bp, self._splitter_bp, cod_agl, descr)
        elif sender is self._tree_dre:
            self._abrir_razao(self._razao_panel_dre, self._splitter_dre, cod_agl, descr)

    def _abrir_razao(
        self, panel: QWidget, splitter: QSplitter,
        cod_cta: str, descricao: str,
    ) -> None:
        """Abre razão da conta no painel inferior do splitter."""
        if self._controller is None:
            return
        razao = self._controller.consultar_razao_completo(cod_cta)

        header = panel._header  # type: ignore[attr-defined]
        tabela = panel._tabela  # type: ignore[attr-defined]
        footer = panel._footer  # type: ignore[attr-defined]
        btn_exp = panel._btn_export  # type: ignore[attr-defined]
        panel._cod_cta_atual = cod_cta  # type: ignore[attr-defined]

        if razao is None or not razao.lancamentos:
            header.setText(
                f"<b>Conta {cod_cta}</b> — {self._html_escape(descricao or '')}"
                "  ·  sem lançamentos no exercício."
            )
            header.setVisible(True)
            tabela.set_rows([])
            footer.setVisible(False)
            btn_exp.setEnabled(False); btn_exp.setVisible(True)
        else:
            header.setText(
                f"<b>Conta {razao.cod_cta}</b> — "
                f"{self._html_escape(razao.descricao or descricao or '')}<br>"
                f"<span style='color:#787A80; font-weight:500'>"
                f"Saldo inicial:</span> R$ {_fmt_brl(razao.saldo_inicial)} "
                f"<b>{razao.ind_dc_inicial}</b>"
            )
            header.setVisible(True)
            tabela.set_rows(
                [self._lancamento_para_dict(l) for l in razao.lancamentos]
            )
            footer.setText(
                f"<span style='color:#787A80'>Totais do exercício:</span> "
                f"D R$ {_fmt_brl(razao.total_debito)}  ·  "
                f"C R$ {_fmt_brl(razao.total_credito)}<br>"
                f"<b>Saldo final:</b> R$ {_fmt_brl(razao.saldo_final)} "
                f"<b>{razao.ind_dc_final}</b>"
            )
            footer.setVisible(True)
            btn_exp.setEnabled(True); btn_exp.setVisible(True)

        # Abre o splitter na proporção 60/40
        sizes = splitter.sizes()
        total = sum(sizes)
        if total > 0:
            splitter.setSizes([int(total * 0.55), int(total * 0.45)])

    # ------------------------------------------------------------
    # Oportunidades de essencialidade (Tema 779) — marcação manual
    # ------------------------------------------------------------

    def _marcar_oportunidade(self) -> None:
        if self._controller is None:
            return
        row = self._tabela_despesas.selected_row()
        if not row:
            return
        cod_cta = row.get("cod_cta", "")
        if not cod_cta:
            return
        from PySide6.QtCore import QSettings
        settings = QSettings("Primetax Solutions", "SpedCrossref")
        marcado_por = settings.value("Parecer/consultor", "", type=str)
        ok = self._controller.marcar_oportunidade(
            cod_cta, marcado_por=marcado_por,
        )
        if ok:
            Toast.show_success(
                self.window(),
                f"Conta {cod_cta} marcada como oportunidade Tema 779.",
            )
            self._recarregar_despesas()
        else:
            Toast.show_error(self.window(), "Falha ao marcar oportunidade.")

    def _desmarcar_oportunidade(self) -> None:
        if self._controller is None:
            return
        row = self._tabela_despesas.selected_row()
        if not row:
            return
        cod_cta = row.get("cod_cta", "")
        if not cod_cta:
            return
        ok = self._controller.desmarcar_oportunidade(cod_cta)
        if ok:
            Toast.show_success(
                self.window(),
                f"Marcação de {cod_cta} removida.",
            )
            self._recarregar_despesas()

    def _recarregar_despesas(self) -> None:
        """Recarrega só a tabela de despesas (após marcar/desmarcar)."""
        if self._controller is None:
            return
        self._linhas_despesas = self._controller.listar_despesas_vs_credito()
        self._tabela_despesas.set_rows(
            [self._despesa_para_dict(d) for d in self._linhas_despesas]
        )
        self._atualizar_contador_oportunidades()

    def _marcar_oportunidade_imob(self) -> None:
        if self._controller is None:
            return
        row = self._tabela_imobilizado.selected_row()
        if not row:
            return
        cod_cta = row.get("cod_cta", "")
        if not cod_cta:
            return
        from PySide6.QtCore import QSettings
        settings = QSettings("Primetax Solutions", "SpedCrossref")
        marcado_por = settings.value("Parecer/consultor", "", type=str)
        ok = self._controller.marcar_oportunidade(
            cod_cta, marcado_por=marcado_por,
        )
        if ok:
            Toast.show_success(
                self.window(),
                f"Conta de imobilizado {cod_cta} marcada como oportunidade.",
            )
            self._recarregar_imobilizado()
        else:
            Toast.show_error(self.window(), "Falha ao marcar oportunidade.")

    def _desmarcar_oportunidade_imob(self) -> None:
        if self._controller is None:
            return
        row = self._tabela_imobilizado.selected_row()
        if not row:
            return
        cod_cta = row.get("cod_cta", "")
        if not cod_cta:
            return
        ok = self._controller.desmarcar_oportunidade(cod_cta)
        if ok:
            Toast.show_success(
                self.window(),
                f"Marcação de {cod_cta} removida.",
            )
            self._recarregar_imobilizado()

    def _recarregar_imobilizado(self) -> None:
        if self._controller is None:
            return
        self._linhas_imobilizado = self._controller.listar_imobilizado_vs_credito()
        self._tabela_imobilizado.set_rows(
            [self._imobilizado_para_dict(d) for d in self._linhas_imobilizado]
        )
        self._atualizar_contador_imobilizado()

    def _atualizar_contador_imobilizado(self) -> None:
        marcadas = [d for d in self._linhas_imobilizado if d.marcada_oportunidade]
        if not marcadas:
            self._lbl_contador_imob.setVisible(False)
            return
        total_saldo = sum((d.saldo_periodo for d in marcadas), Decimal("0"))
        self._lbl_contador_imob.setText(
            f"⚑ <b>{len(marcadas)}</b> conta(s) de imobilizado marcada(s) "
            f"como oportunidade Tema 779 · saldo potencial total: "
            f"<b>R$ {_fmt_brl(total_saldo)}</b>"
        )
        self._lbl_contador_imob.setVisible(True)

    def _atualizar_contador_oportunidades(self) -> None:
        marcadas = [d for d in self._linhas_despesas if d.marcada_oportunidade]
        if not marcadas:
            self._lbl_contador_oport.setVisible(False)
            return
        total_saldo = sum((d.saldo_periodo for d in marcadas), Decimal("0"))
        self._lbl_contador_oport.setText(
            f"⚑ <b>{len(marcadas)}</b> conta(s) marcada(s) como oportunidade "
            f"Tema 779 · saldo potencial total: "
            f"<b>R$ {_fmt_brl(total_saldo)}</b>"
        )
        self._lbl_contador_oport.setVisible(True)

    # ------------------------------------------------------------
    # Despesas × Crédito — drill-down de evidências
    # ------------------------------------------------------------

    def _on_despesa_selecionada(self, row_dict: dict) -> None:
        if self._controller is None:
            return
        cod_cta = row_dict.get("cod_cta", "")
        if not cod_cta:
            return
        descricao = row_dict.get("descricao", "")
        # Habilita botões de ação + ajusta estado conforme já marcada
        self._btn_marcar_oport.setEnabled(True)
        self._btn_desmarcar_oport.setEnabled(True)
        marcada = (row_dict.get("oportunidade") or "").strip() != ""
        self._btn_marcar_oport.setText(
            "⚑ Atualizar marcação Tema 779" if marcada
            else "⚑ Marcar como oportunidade Tema 779"
        )

        # 1. Painel Evidências do crédito (F100/F120/F130)
        evidencias = self._controller.listar_evidencias_credito(cod_cta)
        if not evidencias:
            self._evidencias_header.setText(
                f"<b>Conta {cod_cta}</b> — {self._html_escape(descricao)}<br>"
                "<span style='color:#B23A3A'>"
                "Sem evidências de crédito PIS/COFINS em F100/F120/F130 "
                "(zero crédito tomado). Veja o razão da conta para "
                "reanálise sob Tema 779."
                "</span>"
            )
            self._tabela_evidencias.set_rows([])
        else:
            total = sum((e.valor_base for e in evidencias), Decimal("0"))
            self._evidencias_header.setText(
                f"<b>Conta {cod_cta}</b> — {self._html_escape(descricao)}<br>"
                f"<span style='color:#787A80'>Evidências do crédito:</span> "
                f"{len(evidencias)} linha(s) · base total R$ {_fmt_brl(total)}"
            )
            self._tabela_evidencias.set_rows(
                [self._evidencia_para_dict(e) for e in evidencias]
            )

        # 2. Painel Razão da conta — preenche pra qualquer despesa,
        # com ou sem crédito (auditor pode ver lançamentos contábeis
        # pra fundamentar reanálise sob essencialidade).
        razao = self._controller.consultar_razao_completo(cod_cta)
        panel = self._razao_panel_despesas
        header = panel._header  # type: ignore[attr-defined]
        tabela = panel._tabela  # type: ignore[attr-defined]
        footer = panel._footer  # type: ignore[attr-defined]
        btn_exp = panel._btn_export  # type: ignore[attr-defined]
        panel._cod_cta_atual = cod_cta  # type: ignore[attr-defined]

        if razao is None or not razao.lancamentos:
            header.setText(
                f"<b>Conta {cod_cta}</b> — {self._html_escape(descricao)}"
                "  ·  sem lançamentos no exercício."
            )
            header.setVisible(True)
            tabela.set_rows([])
            footer.setVisible(False)
            btn_exp.setEnabled(False); btn_exp.setVisible(True)
        else:
            header.setText(
                f"<b>Conta {razao.cod_cta}</b> — "
                f"{self._html_escape(razao.descricao or descricao or '')}<br>"
                f"<span style='color:#787A80; font-weight:500'>"
                f"Saldo inicial:</span> R$ {_fmt_brl(razao.saldo_inicial)} "
                f"<b>{razao.ind_dc_inicial}</b>"
            )
            header.setVisible(True)
            tabela.set_rows(
                [self._lancamento_para_dict(l) for l in razao.lancamentos]
            )
            footer.setText(
                f"<span style='color:#787A80'>Totais do exercício:</span> "
                f"D R$ {_fmt_brl(razao.total_debito)}  ·  "
                f"C R$ {_fmt_brl(razao.total_credito)}<br>"
                f"<b>Saldo final:</b> R$ {_fmt_brl(razao.saldo_final)} "
                f"<b>{razao.ind_dc_final}</b>"
            )
            footer.setVisible(True)
            btn_exp.setEnabled(True); btn_exp.setVisible(True)

    def _on_imobilizado_selecionado(self, row_dict: dict) -> None:
        """Seleção numa linha de imobilizado: popula o razão da conta no
        painel inferior e habilita os botões de marcação."""
        if self._controller is None:
            return
        cod_cta = row_dict.get("cod_cta", "")
        if not cod_cta:
            return
        descricao = row_dict.get("descricao", "")
        marcada = (row_dict.get("oportunidade") or "").strip() != ""

        self._btn_marcar_imob.setEnabled(True)
        self._btn_desmarcar_imob.setEnabled(True)
        self._btn_marcar_imob.setText(
            "⚑ Atualizar marcação Tema 779" if marcada
            else "⚑ Marcar como oportunidade Tema 779"
        )

        self._imob_header.setText(
            f"<b>Conta {cod_cta}</b> — {self._html_escape(descricao)}"
            f"{' · <span style=\"color:#008C95\">⚑ marcada</span>' if marcada else ''}"
        )

        # Razão da conta no painel inferior
        razao = self._controller.consultar_razao_completo(cod_cta)
        panel = self._razao_panel_imob
        header = panel._header  # type: ignore[attr-defined]
        tabela = panel._tabela  # type: ignore[attr-defined]
        footer = panel._footer  # type: ignore[attr-defined]
        btn_exp = panel._btn_export  # type: ignore[attr-defined]
        panel._cod_cta_atual = cod_cta  # type: ignore[attr-defined]

        if razao is None or not razao.lancamentos:
            header.setText(
                f"<b>Conta {cod_cta}</b> — sem lançamentos no exercício."
            )
            header.setVisible(True)
            tabela.set_rows([])
            footer.setVisible(False)
            btn_exp.setEnabled(False); btn_exp.setVisible(True)
            return

        header.setText(
            f"<b>Conta {razao.cod_cta}</b> — "
            f"{self._html_escape(razao.descricao or descricao or '')}<br>"
            f"<span style='color:#787A80; font-weight:500'>"
            f"Saldo inicial:</span> R$ {_fmt_brl(razao.saldo_inicial)} "
            f"<b>{razao.ind_dc_inicial}</b>"
        )
        header.setVisible(True)
        tabela.set_rows(
            [self._lancamento_para_dict(l) for l in razao.lancamentos]
        )
        footer.setText(
            f"<span style='color:#787A80'>Totais do exercício:</span> "
            f"D R$ {_fmt_brl(razao.total_debito)}  ·  "
            f"C R$ {_fmt_brl(razao.total_credito)}<br>"
            f"<b>Saldo final:</b> R$ {_fmt_brl(razao.saldo_final)} "
            f"<b>{razao.ind_dc_final}</b>"
        )
        footer.setVisible(True)
        btn_exp.setEnabled(True); btn_exp.setVisible(True)

    # ------------------------------------------------------------
    # Conversões
    # ------------------------------------------------------------

    @staticmethod
    def _lancamento_para_dict(l: LancamentoRazao) -> dict:
        return {
            "data": _fmt_data_brl(l.data),
            "num_lcto": l.num_lcto,
            "sub_conta": l.sub_conta,
            "historico": l.historico,
            "debito": l.debito,
            "credito": l.credito,
            "contrapartida": (
                f"{l.contrapartida_cta} · {l.contrapartida_descr}"
                if l.contrapartida_cta else ""
            ),
            "saldo": l.saldo_corrente,
        }

    @staticmethod
    def _despesa_para_dict(d: DespesaVsCredito) -> dict:
        return {
            "oportunidade": "⚑" if d.marcada_oportunidade else "",
            "cod_cta": d.cod_cta,
            "descricao": d.descricao,
            "saldo_periodo": d.saldo_periodo,
            "credito_pis_cofins": d.credito_pis_cofins,
            "tem_credito_badge": (
                BadgeStatus.OK if d.tem_credito else BadgeStatus.ALTO
            ),
        }

    @staticmethod
    def _evidencia_para_dict(e: EvidenciaCredito) -> dict:
        return {
            "registro": e.registro,
            "ano_mes": _fmt_periodo_brl(e.ano_mes),
            "valor_base": e.valor_base,
            "nat_bc_cred": e.nat_bc_cred,
            "arquivo_linha": (
                f"{Path(e.arquivo).name}:{e.linha_arquivo}"
                if e.arquivo else f"L{e.linha_arquivo}"
            ),
            "descricao": e.descricao,
        }

    # ------------------------------------------------------------
    # Colunas
    # ------------------------------------------------------------

    @staticmethod
    def _cols_razao() -> list[ColumnSpec]:
        return [
            ColumnSpec(id="data", header="Data", kind="text", width=100),
            ColumnSpec(id="num_lcto", header="Nº Lcto", kind="text", width=80),
            ColumnSpec(id="sub_conta", header="Sub-conta", kind="text", width=110),
            ColumnSpec(id="historico", header="Histórico", kind="text", width=300),
            ColumnSpec(id="contrapartida", header="Contrapartida", kind="text", width=240),
            ColumnSpec(id="debito", header="Débito", kind="money", width=130),
            ColumnSpec(id="credito", header="Crédito", kind="money", width=130),
            ColumnSpec(id="saldo", header="Saldo", kind="money", width=140),
        ]

    @staticmethod
    def _cols_imobilizado() -> list[ColumnSpec]:
        return [
            ColumnSpec(id="oportunidade", header="⚑", kind="text", width=46),
            ColumnSpec(id="cod_cta", header="Conta", kind="text", width=120),
            ColumnSpec(id="saldo_periodo", header="Saldo no AC", kind="money", width=160),
            ColumnSpec(id="credito_f120", header="Crédito F120 (depr.)", kind="money", width=160),
            ColumnSpec(id="credito_f130", header="Crédito F130 (aquis.)", kind="money", width=160),
            ColumnSpec(id="tem_credito_badge", header="Status", kind="badge", width=110),
            ColumnSpec(id="descricao", header="Descrição da conta", kind="text", width=360),
        ]

    @staticmethod
    def _imobilizado_para_dict(d: ImobilizadoVsCredito) -> dict:
        return {
            "oportunidade": "⚑" if d.marcada_oportunidade else "",
            "cod_cta": d.cod_cta,
            "descricao": d.descricao,
            "saldo_periodo": d.saldo_periodo,
            "credito_f120": d.credito_f120,
            "credito_f130": d.credito_f130,
            "tem_credito_badge": (
                BadgeStatus.OK if d.tem_credito else BadgeStatus.ALTO
            ),
        }

    @staticmethod
    def _cols_despesas() -> list[ColumnSpec]:
        return [
            ColumnSpec(id="oportunidade", header="⚑", kind="text", width=46),
            ColumnSpec(id="cod_cta", header="Conta", kind="text", width=120),
            ColumnSpec(id="saldo_periodo", header="Saldo no AC", kind="money", width=160),
            ColumnSpec(id="credito_pis_cofins", header="Crédito PIS/COFINS", kind="money", width=180),
            ColumnSpec(id="tem_credito_badge", header="Status", kind="badge", width=110),
            ColumnSpec(id="descricao", header="Descrição da conta", kind="text", width=400),
        ]

    @staticmethod
    def _cols_evidencias() -> list[ColumnSpec]:
        return [
            ColumnSpec(id="registro", header="Registro", kind="text", width=90),
            ColumnSpec(id="ano_mes", header="Período", kind="text", width=100),
            ColumnSpec(id="valor_base", header="Valor base", kind="money", width=140),
            ColumnSpec(id="nat_bc_cred", header="NAT_BC_CRED", kind="text", width=130),
            ColumnSpec(id="arquivo_linha", header="Arquivo:Linha SPED", kind="text", width=320),
            ColumnSpec(id="descricao", header="Descrição", kind="text", width=320),
        ]

    # ------------------------------------------------------------
    # Exports individualizados por aba
    # ------------------------------------------------------------

    def _exportar_bp(self) -> None:
        if not self._linhas_bp or self._cliente is None:
            Toast.show_warning(self.window(), "Sem dados de BP para exportar.")
            return
        destino = self._destino_xlsx("balanco_patrimonial")

        def _salvar():
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Balanço Patrimonial"
            self._aba_balanco_excel(ws, self._linhas_bp, modo="bp")
            wb.save(str(destino))

        if self._executar_export("Balanço Patrimonial", _salvar):
            self._toast_export_sucesso(destino, "Balanço Patrimonial")
            self.csv_exportado.emit(destino)

    def _exportar_dre(self) -> None:
        if not self._linhas_dre or self._cliente is None:
            Toast.show_warning(self.window(), "Sem dados de DRE para exportar.")
            return
        destino = self._destino_xlsx("dre")

        def _salvar():
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "DRE"
            self._aba_balanco_excel(ws, self._linhas_dre, modo="dre")
            wb.save(str(destino))

        if self._executar_export("DRE", _salvar):
            self._toast_export_sucesso(destino, "DRE")
            self.csv_exportado.emit(destino)

    def _exportar_despesas(self) -> None:
        if not self._linhas_despesas or self._cliente is None:
            Toast.show_warning(self.window(), "Sem dados de despesas para exportar.")
            return
        destino = self._destino_xlsx("despesas_vs_credito")

        def _salvar():
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Despesas × Crédito"
            self._aba_despesas_excel(ws, self._linhas_despesas)
            wb.save(str(destino))

        if self._executar_export("Despesas × Crédito", _salvar):
            self._toast_export_sucesso(destino, "Despesas × Crédito")
            self.csv_exportado.emit(destino)

    def _exportar_razao_atual(self) -> None:
        """Exporta razão da conta atual (do painel ativo nas abas BP, DRE,
        Despesas × Crédito ou Imobilizado × Crédito)."""
        ativo_idx = self._tabs.currentIndex()
        if ativo_idx == 0:
            panel = self._razao_panel_bp
        elif ativo_idx == 1:
            panel = self._razao_panel_dre
        elif ativo_idx == 2:
            panel = self._razao_panel_despesas
        elif ativo_idx == 3:
            panel = self._razao_panel_imob
        else:
            return
        cod_cta = getattr(panel, "_cod_cta_atual", "")
        if not cod_cta or self._controller is None or self._cliente is None:
            return
        razao = self._controller.consultar_razao_completo(cod_cta)
        if razao is None:
            Toast.show_warning(self.window(), "Sem razão pra exportar.")
            return
        destino = self._destino_xlsx(f"razao_{cod_cta.replace('.', '_')}")

        def _salvar():
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Razão {cod_cta}"[:31]
            self._aba_razao_excel(ws, razao)
            wb.save(str(destino))

        if self._executar_export(f"Razão da conta {cod_cta}", _salvar):
            self._toast_export_sucesso(destino, f"Razão da conta {cod_cta}")
            self.csv_exportado.emit(destino)

    def _destino_xlsx(self, prefixo: str) -> Path:
        """Caminho do Excel exportado: pasta Downloads do usuário,
        com nome contendo cliente + ano para evitar colisão entre
        exports de clientes/períodos diferentes.

        Ex: ~/Downloads/primetax_balanco_patrimonial_TELLAIO_TEXTIL_LTDA_2018.xlsx
        """
        assert self._cliente is not None
        downloads = QStandardPaths.writableLocation(
            QStandardPaths.StandardLocation.DownloadLocation
        )
        # Fallback caso QStandardPaths retorne vazio (raro mas possível)
        base = Path(downloads) if downloads else Path.home() / "Downloads"
        base.mkdir(parents=True, exist_ok=True)

        # Slug do cliente — letras/números/underscore para nome de arquivo seguro
        slug = "".join(
            c if c.isalnum() else "_"
            for c in self._cliente.razao_social.upper()
        ).strip("_")
        # Reduz múltiplos underscores
        while "__" in slug:
            slug = slug.replace("__", "_")

        nome = f"primetax_{prefixo}_{slug}_{self._cliente.ano_calendario}.xlsx"
        return base / nome

    def _executar_export(self, descricao: str, save_func) -> bool:
        """Executa save_func com feedback visual modal claro:
          - QProgressDialog modal "Exportando ..." (impossível de ignorar)
          - Cursor wait reforça
          - Toast vermelho de erro se save_func falhar
          - Toast verde de sucesso (chamado pelo caller após esta função)
        Retorna True em sucesso, False em erro.
        Sem QThread porque Excel costuma salvar em <1s; o diálogo modal
        + processEvents() garante que o usuário vê o feedback.
        """
        dialog = QProgressDialog(
            f"Exportando {descricao}...\n\nAguarde alguns instantes.",
            "",  # sem botão cancelar (texto vazio + setCancelButton(None))
            0, 0,  # range 0,0 = barra indefinida (animação contínua)
            self.window(),
        )
        dialog.setWindowTitle("Primetax SPED — Exportando")
        dialog.setCancelButton(None)
        dialog.setWindowModality(Qt.WindowModality.ApplicationModal)
        dialog.setMinimumDuration(0)  # mostra imediatamente, sem delay
        dialog.setMinimumWidth(360)
        dialog.setStyleSheet(
            "QProgressDialog { background: #FFFFFF; }"
            "QProgressDialog QLabel { color: #53565A; font-size: 10pt; padding: 8px; }"
            "QProgressBar { border: 1px solid #D1D3D6; border-radius: 2px;"
            "               background: #F7F7F8; height: 8px; }"
            "QProgressBar::chunk { background: #008C95; }"
        )
        dialog.show()
        # Força renderização do diálogo ANTES do save bloquear
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        QApplication.processEvents()
        QApplication.processEvents()  # 2x pra garantir que pinta
        try:
            save_func()
            return True
        except Exception as exc:  # noqa: BLE001
            logger.exception("falha em export %s", descricao)
            Toast.show_error(
                self.window(),
                f"Falha ao exportar {descricao}: {exc}",
            )
            return False
        finally:
            dialog.close()
            QApplication.restoreOverrideCursor()

    def _toast_export_sucesso(self, destino: Path, descricao: str) -> None:
        """Toast verde com duração maior + botão 'Abrir pasta' do arquivo
        gerado. Mostra ao auditor onde o Excel foi salvo (data/output/...)."""
        Toast.show_success(
            self.window(),
            f"{descricao} salvo em: {destino.parent}\\{destino.name}",
            duration_ms=8000,
            action=ToastAction(
                label="Abrir pasta",
                callback=lambda p=destino.parent: QDesktopServices.openUrl(
                    QUrl.fromLocalFile(str(p))
                ),
            ),
        )

    def _aba_balanco_excel(self, ws, linhas: list[LinhaBalanco], *, modo: str) -> None:
        from openpyxl.styles import Font, PatternFill
        headers = ["Cód. Agl.", "Nível", "Tipo", "Descrição", "Valor"]
        for c, h in enumerate(headers, start=1):
            cel = ws.cell(row=1, column=c, value=h)
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="008C95")
        for r, l in enumerate(linhas, start=2):
            valor = self._valor_signed(l, modo=modo)
            ws.cell(row=r, column=1, value=l.cod_agl)
            ws.cell(row=r, column=2, value=l.nivel)
            ws.cell(row=r, column=3, value="Sintética" if (l.ind_cod_agl or "").upper() == "T" else "Analítica")
            ws.cell(row=r, column=4, value=("    " * max(0, l.nivel - 1)) + l.descricao)
            ws.cell(row=r, column=5, value=float(valor))

    def _aba_despesas_excel(self, ws, linhas: list[DespesaVsCredito]) -> None:
        from openpyxl.styles import Font, PatternFill
        headers = ["Oportunidade Tema 779", "Conta", "Descrição",
                   "Saldo no AC", "Crédito PIS/COFINS", "Tem crédito?",
                   "Nota do auditor"]
        for c, h in enumerate(headers, start=1):
            cel = ws.cell(row=1, column=c, value=h)
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="008C95")
        for r, d in enumerate(linhas, start=2):
            ws.cell(row=r, column=1, value="Sim" if d.marcada_oportunidade else "")
            ws.cell(row=r, column=2, value=d.cod_cta)
            ws.cell(row=r, column=3, value=d.descricao)
            ws.cell(row=r, column=4, value=float(d.saldo_periodo))
            ws.cell(row=r, column=5, value=float(d.credito_pis_cofins))
            ws.cell(row=r, column=6, value="Sim" if d.tem_credito else "Não")
            ws.cell(row=r, column=7, value=d.nota_oportunidade)

    def _exportar_imobilizado(self) -> None:
        if not self._linhas_imobilizado or self._cliente is None:
            Toast.show_warning(
                self.window(),
                "Sem contas de imobilizado para exportar.",
            )
            return
        destino = self._destino_xlsx("imobilizado_vs_credito")

        def _salvar():
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            wb = Workbook()
            ws = wb.active
            ws.title = "Imobilizado × Crédito"
            headers = [
                "Oportunidade Tema 779", "Conta", "Descrição", "Saldo no AC",
                "Crédito F120 (depreciação)", "Crédito F130 (aquisição)",
                "Tem crédito?", "Nota do auditor",
            ]
            for c, h in enumerate(headers, start=1):
                cel = ws.cell(row=1, column=c, value=h)
                cel.font = Font(bold=True, color="FFFFFF")
                cel.fill = PatternFill("solid", fgColor="008C95")
            for r, d in enumerate(self._linhas_imobilizado, start=2):
                ws.cell(row=r, column=1, value="Sim" if d.marcada_oportunidade else "")
                ws.cell(row=r, column=2, value=d.cod_cta)
                ws.cell(row=r, column=3, value=d.descricao)
                ws.cell(row=r, column=4, value=float(d.saldo_periodo))
                ws.cell(row=r, column=5, value=float(d.credito_f120))
                ws.cell(row=r, column=6, value=float(d.credito_f130))
                ws.cell(row=r, column=7, value="Sim" if d.tem_credito else "Não")
                ws.cell(row=r, column=8, value=d.nota_oportunidade)
            wb.save(str(destino))

        if self._executar_export("Imobilizado × Crédito", _salvar):
            self._toast_export_sucesso(destino, "Imobilizado × Crédito")
            self.csv_exportado.emit(destino)

    def _exportar_relatorio_imobilizado(self) -> None:
        """Relatório só com contas de imobilizado marcadas como oportunidade."""
        if self._cliente is None:
            return
        marcadas = [d for d in self._linhas_imobilizado if d.marcada_oportunidade]
        if not marcadas:
            Toast.show_warning(
                self.window(),
                "Nenhuma conta de imobilizado marcada como oportunidade. "
                "Selecione uma conta e clique em '⚑ Marcar como oportunidade'.",
            )
            return
        destino = self._destino_xlsx("relatorio_oportunidades_imobilizado")

        def _salvar():
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = Workbook()
            ws = wb.active
            ws.title = "Oportunidades Imobilizado"

            ws.cell(
                row=1, column=1,
                value="RELATÓRIO DE OPORTUNIDADES — IMOBILIZADO PIS/COFINS",
            ).font = Font(bold=True, size=14, color="008C95")
            ws.cell(
                row=2, column=1,
                value=f"Cliente: {self._cliente.razao_social}  ·  "
                      f"CNPJ: {self._cliente.cnpj_formatado()}  ·  "
                      f"AC {self._cliente.ano_calendario}",
            ).font = Font(italic=True, size=10)
            ws.cell(
                row=3, column=1,
                value="Base: art. 3º, VI da Lei 10.637/2002 e 10.833/2003 "
                      "(crédito sobre bens do imobilizado adquiridos para "
                      "uso na produção)",
            ).font = Font(size=9, color="787A80")

            headers = ["Conta", "Descrição", "Saldo no AC (R$)",
                       "Crédito F120 atual", "Crédito F130 atual",
                       "Nota do auditor"]
            for c, h in enumerate(headers, start=1):
                cel = ws.cell(row=5, column=c, value=h)
                cel.font = Font(bold=True, color="FFFFFF")
                cel.fill = PatternFill("solid", fgColor="008C95")
                cel.alignment = Alignment(horizontal="center")

            total_saldo = Decimal("0")
            r = 6
            for d in marcadas:
                ws.cell(row=r, column=1, value=d.cod_cta)
                ws.cell(row=r, column=2, value=d.descricao)
                ws.cell(row=r, column=3, value=float(d.saldo_periodo))
                ws.cell(row=r, column=4, value=float(d.credito_f120))
                ws.cell(row=r, column=5, value=float(d.credito_f130))
                ws.cell(row=r, column=6, value=d.nota_oportunidade)
                total_saldo += d.saldo_periodo
                r += 1

            ws.cell(row=r + 1, column=1, value="TOTAL").font = Font(bold=True)
            ws.cell(row=r + 1, column=3, value=float(total_saldo)).font = Font(bold=True)

            wb.save(str(destino))

        if self._executar_export(
            "Relatório de oportunidades — Imobilizado", _salvar,
        ):
            self._toast_export_sucesso(
                destino, "Relatório de oportunidades — Imobilizado"
            )
            self.csv_exportado.emit(destino)

    def _exportar_relatorio_oportunidades(self) -> None:
        """Exporta apenas as despesas marcadas como oportunidade Tema 779
        com saldo potencial e descrição. Pra entrega ao cliente / parecer."""
        if self._cliente is None:
            return
        marcadas = [d for d in self._linhas_despesas if d.marcada_oportunidade]
        if not marcadas:
            Toast.show_warning(
                self.window(),
                "Nenhuma despesa marcada como oportunidade. "
                "Selecione uma despesa e clique em '⚑ Marcar como oportunidade'.",
            )
            return
        destino = self._destino_xlsx("relatorio_oportunidades_tema_779")

        def _salvar():
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = Workbook()
            ws = wb.active
            ws.title = "Oportunidades Tema 779"

            # Cabeçalho institucional
            ws.cell(row=1, column=1, value="RELATÓRIO DE OPORTUNIDADES — ESSENCIALIDADE TEMA 779").font = Font(
                bold=True, size=14, color="008C95",
            )
            ws.cell(
                row=2, column=1,
                value=f"Cliente: {self._cliente.razao_social}  ·  "
                      f"CNPJ: {self._cliente.cnpj_formatado()}  ·  "
                      f"AC {self._cliente.ano_calendario}",
            ).font = Font(italic=True, size=10)
            ws.cell(
                row=3, column=1,
                value="Base: REsp 1.221.170/PR (Tema 779) — essencialidade "
                      "e relevância como critérios de creditamento PIS/COFINS",
            ).font = Font(size=9, color="787A80")

            # Cabeçalho da tabela
            headers = ["Conta", "Descrição da conta", "Saldo no AC (R$)",
                       "Crédito PIS/COFINS atual (R$)", "Marcado em",
                       "Marcado por", "Nota do auditor"]
            for c, h in enumerate(headers, start=1):
                cel = ws.cell(row=5, column=c, value=h)
                cel.font = Font(bold=True, color="FFFFFF")
                cel.fill = PatternFill("solid", fgColor="008C95")
                cel.alignment = Alignment(horizontal="center")

            total_saldo = Decimal("0")
            r = 6
            for d in marcadas:
                ws.cell(row=r, column=1, value=d.cod_cta)
                ws.cell(row=r, column=2, value=d.descricao)
                ws.cell(row=r, column=3, value=float(d.saldo_periodo))
                ws.cell(row=r, column=4, value=float(d.credito_pis_cofins))
                ws.cell(row=r, column=5, value="—")  # marcado_em não vem na view
                ws.cell(row=r, column=6, value="—")
                ws.cell(row=r, column=7, value=d.nota_oportunidade)
                total_saldo += d.saldo_periodo
                r += 1

            # Linha total
            ws.cell(row=r + 1, column=1, value="TOTAL").font = Font(bold=True)
            ws.cell(row=r + 1, column=3, value=float(total_saldo)).font = Font(bold=True)

            wb.save(str(destino))

        if self._executar_export(
            "Relatório de oportunidades — Tema 779", _salvar,
        ):
            self._toast_export_sucesso(
                destino, "Relatório de oportunidades — Tema 779"
            )
            self.csv_exportado.emit(destino)

    def _aba_razao_excel(self, ws, razao: RazaoConta) -> None:
        from openpyxl.styles import Font, PatternFill
        ws.cell(row=1, column=1, value=f"Razão da conta {razao.cod_cta}").font = Font(
            bold=True, size=12,
        )
        ws.cell(row=2, column=1, value=razao.descricao).font = Font(italic=True)
        ws.cell(
            row=3, column=1,
            value=f"Saldo inicial: R$ {_fmt_brl(razao.saldo_inicial)} {razao.ind_dc_inicial}",
        )
        headers = ["Data", "Nº Lcto", "Sub-conta", "Histórico",
                   "Contrapartida", "Débito", "Crédito", "Saldo corrente"]
        for c, h in enumerate(headers, start=1):
            cel = ws.cell(row=5, column=c, value=h)
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="008C95")
        r = 6
        for l in razao.lancamentos:
            ws.cell(row=r, column=1, value=_fmt_data_brl(l.data))
            ws.cell(row=r, column=2, value=l.num_lcto)
            ws.cell(row=r, column=3, value=l.sub_conta)
            ws.cell(row=r, column=4, value=l.historico)
            contrapartida = (
                f"{l.contrapartida_cta} · {l.contrapartida_descr}"
                if l.contrapartida_cta else ""
            )
            ws.cell(row=r, column=5, value=contrapartida)
            ws.cell(row=r, column=6, value=float(l.debito))
            ws.cell(row=r, column=7, value=float(l.credito))
            ws.cell(row=r, column=8, value=float(l.saldo_corrente))
            r += 1
        # Footer com totais
        ws.cell(row=r + 1, column=1, value="Totais").font = Font(bold=True)
        ws.cell(row=r + 1, column=6, value=float(razao.total_debito)).font = Font(bold=True)
        ws.cell(row=r + 1, column=7, value=float(razao.total_credito)).font = Font(bold=True)
        ws.cell(
            row=r + 2, column=1,
            value=f"Saldo final: R$ {_fmt_brl(razao.saldo_final)} {razao.ind_dc_final}",
        ).font = Font(bold=True)

    # ------------------------------------------------------------
    # Estilos
    # ------------------------------------------------------------

    @staticmethod
    def _qss_tabs() -> str:
        return f"""
        QTabWidget::pane {{
            border: 1px solid #D1D3D6;
            border-radius: 2px;
            background: #FFFFFF;
            top: -1px;
        }}
        QTabBar::tab {{
            background: #F0F0F1;
            color: #53565A;
            border: 1px solid #D1D3D6;
            padding: 6px 14px;
            font-size: 10pt;
            font-weight: 500;
            margin-right: 2px;
        }}
        QTabBar::tab:selected {{
            background: #FFFFFF;
            color: {_PRIMARY_COLOR};
            border-bottom: 1px solid #FFFFFF;
            font-weight: 600;
        }}
        QTabBar::tab:hover:!selected {{
            background: #E6F3F4;
        }}
        """

    @staticmethod
    def _qss_tree() -> str:
        return """
        QTreeView {
            background: #FFFFFF;
            alternate-background-color: #F7F7F8;
            color: #53565A;
            font-size: 10pt;
            border: 1px solid #D1D3D6;
            selection-background-color: #E6F3F4;
            selection-color: #53565A;
        }
        QTreeView::item {
            padding: 4px 6px;
        }
        QTreeView::item:selected {
            background: #E6F3F4;
            color: #53565A;
        }
        QHeaderView::section {
            background: #008C95;
            color: #FFFFFF;
            font-weight: 600;
            padding: 6px 8px;
            border: 0;
            border-right: 1px solid #006F76;
        }
        """

    @staticmethod
    def _qss_btn_primario() -> str:
        return f"""
        QPushButton {{
            background: {_PRIMARY_COLOR}; color: #FFFFFF; border: none;
            border-radius: 2px; padding: 6px 14px;
            font-size: 10pt; font-weight: 500;
        }}
        QPushButton:hover {{ background: #00A4AE; }}
        QPushButton:pressed {{ background: #006F76; }}
        QPushButton:disabled {{ background: #B3D7DA; color: #FFFFFF; }}
        """

    @staticmethod
    def _qss_btn_secundario() -> str:
        return f"""
        QPushButton {{
            background: #FFFFFF; color: {_PRIMARY_COLOR};
            border: 1px solid {_PRIMARY_COLOR};
            border-radius: 2px; padding: 6px 14px;
            font-size: 10pt; font-weight: 500;
        }}
        QPushButton:hover {{ background: #E6F3F4; }}
        QPushButton:disabled {{ color: #B3D7DA; border-color: #B3D7DA; }}
        """

    @staticmethod
    def _html_escape(s: str) -> str:
        return (
            (s or "")
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )

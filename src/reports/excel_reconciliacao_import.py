"""
Importador do template preenchido de reconciliação de plano de contas (CLAUDE.md §16.6).

Comando CLI: `primetax-sped reconciliacao-import <cnpj> <ano> <arquivo.xlsx>`

Lê a aba 'Plano de Contas (atual)' do Excel gerado por `reconciliacao-template`,
extrai as linhas em que o auditor preencheu a coluna 'COD_CTA antigo' e grava
os mapeamentos na tabela `reconciliacao_override`. Mapeamentos já existentes
para o mesmo (cnpj, ano, cod_cta_atual) são atualizados via UPSERT.

A importação não altera dados dos arquivos SPED originais (§4 princípio 2) —
apenas adiciona uma camada paralela de reconciliação manual que é consultada
pelos cruzamentos em modo degradado.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from src.db.repo import Repositorio

# Nome da aba esperada no template (criada por excel_reconciliacao.gerar)
_ABA_PLANO_ATUAL = "Plano de Contas (atual)"

# Posições esperadas das colunas (1-indexed no openpyxl)
_COL_COD_CTA_ATUAL = 1
_COL_COD_CTA_ANTIGO = 8
_COL_NOME_ANTIGO = 9
_COL_OBSERVACOES = 10


@dataclass
class ResultadoImport:
    """Sumário da importação de overrides."""
    arquivo: str
    cnpj: str
    ano_calendario: int
    linhas_importadas: int
    linhas_ignoradas: int


def importar_template(
    repo: Repositorio,
    ano_calendario: int,
    caminho_xlsx: Path,
    *,
    substituir: bool = True,
) -> ResultadoImport:
    """
    Importa o template preenchido e grava os mapeamentos como overrides.

    Args:
        repo: Repositório do CNPJ × ano-calendário.
        ano_calendario: Ano-calendário do diagnóstico.
        caminho_xlsx: Caminho do Excel preenchido.
        substituir: Se True (padrão), limpa overrides anteriores antes de inserir
            (evita acúmulo de mapeamentos obsoletos de importações passadas).

    Returns:
        ResultadoImport com contagens.

    Raises:
        ValueError: se o arquivo não existir, não tiver a aba esperada,
            ou nenhuma linha tiver COD_CTA antigo preenchido.
    """
    if not caminho_xlsx.exists():
        raise ValueError(f"Arquivo não encontrado: {caminho_xlsx}")

    try:
        wb = load_workbook(str(caminho_xlsx), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Arquivo inválido ou corrompido: {exc}") from exc

    if _ABA_PLANO_ATUAL not in wb.sheetnames:
        raise ValueError(
            f"Aba '{_ABA_PLANO_ATUAL}' não encontrada no arquivo. "
            f"Abas disponíveis: {wb.sheetnames}. "
            f"Gere o template via `reconciliacao-template` antes de preencher."
        )

    ws = wb[_ABA_PLANO_ATUAL]

    mapeamentos: list[tuple[str, str, str, str]] = []
    linhas_ignoradas = 0

    # Linha 1 é cabeçalho; dados começam na linha 2
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue
        cod_atual = (row[_COL_COD_CTA_ATUAL - 1] or "")
        cod_antigo = (row[_COL_COD_CTA_ANTIGO - 1] or "")
        nome_antigo = (row[_COL_NOME_ANTIGO - 1] or "") if len(row) >= _COL_NOME_ANTIGO else ""
        observacoes = (row[_COL_OBSERVACOES - 1] or "") if len(row) >= _COL_OBSERVACOES else ""

        cod_atual_str = str(cod_atual).strip()
        cod_antigo_str = str(cod_antigo).strip()

        if not cod_atual_str:
            continue  # linha em branco no meio do plano
        if not cod_antigo_str:
            linhas_ignoradas += 1
            continue  # auditor não preencheu — conta não mudou

        mapeamentos.append((
            cod_atual_str,
            cod_antigo_str,
            str(nome_antigo).strip(),
            str(observacoes).strip(),
        ))

    wb.close()

    if not mapeamentos:
        raise ValueError(
            "Nenhuma linha com 'COD_CTA antigo' preenchida. "
            "Preencha a coluna em amarelo antes de importar."
        )

    cnpj = repo.cnpj
    conn = repo.conexao()
    try:
        with conn:
            if substituir:
                repo.limpar_reconciliacao_overrides(conn, cnpj, ano_calendario)
            for cod_atual, cod_antigo, nome, obs in mapeamentos:
                repo.inserir_reconciliacao_override(
                    conn,
                    cnpj=cnpj,
                    ano_calendario=ano_calendario,
                    cod_cta_atual=cod_atual,
                    cod_cta_antigo=cod_antigo,
                    nome_antigo=nome,
                    observacoes=obs,
                    arquivo_origem=str(caminho_xlsx.resolve()),
                )
    finally:
        conn.close()

    return ResultadoImport(
        arquivo=str(caminho_xlsx),
        cnpj=cnpj,
        ano_calendario=ano_calendario,
        linhas_importadas=len(mapeamentos),
        linhas_ignoradas=linhas_ignoradas,
    )

"""
Gerador do template de reconciliação manual de plano de contas (CLAUDE.md §16.6).

Comando CLI: `primetax-sped reconciliacao-template --cnpj X --ano 2025`

Saída: Excel com o plano de contas analítico da ECD (I050) em uma aba, com
colunas adicionais em branco para o auditor preencher manualmente a
correspondência entre o plano pré-mudança e o plano pós-mudança.

Uso previsto:
  - ECDs com IND_MUDANC_PC='1' onde o Bloco C (C050/C052/C155) está ausente
    ou incompleto, classificadas pelo sistema como "suspeita" ou "ausente".
  - Auditor preenche a coluna "COD_CTA antigo" para cada conta que sofreu
    reclassificação no período.
  - Versões futuras do sistema poderão aceitar o arquivo preenchido de volta
    via `primetax-sped reconciliacao-import` (fora do escopo atual, §16.6).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.db.repo import Repositorio

_TEAL = "008C95"
_CINZA = "53565A"
_BRANCO = "FFFFFF"
_AMARELO_CLARO = "FFF9C4"

_FILL_TEAL = PatternFill("solid", fgColor=_TEAL)
_FILL_CINZA = PatternFill("solid", fgColor=_CINZA)
_FILL_AMARELO = PatternFill("solid", fgColor=_AMARELO_CLARO)

_FONT_HEADER = Font(bold=True, color=_BRANCO, name="Calibri", size=11)
_FONT_TITULO = Font(bold=True, color=_CINZA, name="Calibri", size=13)
_FONT_NORMAL = Font(name="Calibri", size=10)
_FONT_INSTRUCAO = Font(italic=True, color=_CINZA, name="Calibri", size=10)

_ROTULO_NATUREZA = {
    "01": "Ativo",
    "02": "Passivo",
    "03": "Patrimônio Líquido",
    "04": "Resultado",
    "05": "Compensação",
    "09": "Outras",
}


def gerar(
    repo: Repositorio,
    ano_calendario: int,
    destino: Path,
) -> Path:
    """
    Gera o template de reconciliação manual para o CNPJ × ano-calendário.

    Returns:
        Caminho do arquivo .xlsx gerado.

    Raises:
        ValueError: se não há plano de contas I050 importado para o CNPJ × ano.
    """
    wb = Workbook()
    wb.remove(wb.active)

    conn = repo.conexao()
    try:
        i050 = repo.consultar_ecd_i050(conn, repo.cnpj, ano_calendario)
        c050 = repo.consultar_ecd_c050(conn, repo.cnpj, ano_calendario)
        ind_mudanc_pc = repo.consultar_ecd_ind_mudanc_pc(conn, repo.cnpj, ano_calendario)
    finally:
        conn.close()

    if not i050:
        raise ValueError(
            f"Sem plano de contas I050 para CNPJ={repo.cnpj} AC={ano_calendario}."
            " Importe a ECD antes de gerar o template."
        )

    _aba_instrucoes(wb, repo.cnpj, ano_calendario, ind_mudanc_pc, len(i050), len(c050))
    _aba_plano_contas(wb, i050)
    if c050:
        _aba_plano_antigo(wb, c050)

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(destino))
    return destino


def _aba_instrucoes(
    wb, cnpj: str, ano: int, ind_mudanc_pc: str | None,
    qtd_contas_atuais: int, qtd_contas_antigas: int,
) -> None:
    ws = wb.create_sheet("Instruções")

    ws.cell(row=1, column=1, value="Template de Reconciliação Manual — Plano de Contas").font = _FONT_TITULO
    ws.cell(row=2, column=1, value=f"CNPJ: {cnpj}  /  Ano-calendário: {ano}").font = _FONT_NORMAL
    ws.cell(
        row=3, column=1,
        value=(
            f"IND_MUDANC_PC = '{ind_mudanc_pc or '—'}'  |  "
            f"Plano atual (I050): {qtd_contas_atuais} contas  |  "
            f"Plano antigo (C050): {qtd_contas_antigas} contas"
        ),
    ).font = _FONT_NORMAL

    tem_c050 = qtd_contas_antigas > 0
    instrucoes = [
        "",
        "Este arquivo é gerado pelo comando:",
        "   primetax-sped reconciliacao-template {cnpj} {ano}",
        "",
        "Uso previsto:",
        "  1. A ECD deste CNPJ × ano foi classificada como suspeita ou ausente",
        "     para reconciliação de plano de contas (CLAUDE.md §16.2) — o que",
        "     significa que houve mudança declarada mas o Bloco C (registros",
        "     C050/C052/C155 da ECD) está vazio ou incompleto.",
        "",
        "  2. Sem reconciliação, cruzamentos dependentes de COD_CTA específica",
        "     (CR-40, CR-41, CR-44) não podem ser executados em modo integral.",
        "     Cruzamentos com suporte a modo degradado (CR-38, CR-39, CR-43)",
        "     são rebaixados para agregação por COD_NAT (§16.3).",
        "",
        "  3. Na aba 'Plano de Contas (atual)', preencha, para cada conta que",
        "     sofreu reclassificação no período, as colunas em amarelo:",
        "       - COD_CTA antigo: código da conta ANTES da mudança",
        "       - NOME antigo: nome da conta antiga (opcional, para conferência)",
        "       - Observações: notas do auditor (opcional)",
        "",
        "  4. Contas que não mudaram não precisam de preenchimento.",
        "",
    ]
    if tem_c050:
        instrucoes.extend([
            "  5. A aba 'Plano Antigo (C050)' lista o plano recuperado da ECD",
            "     anterior, para consulta durante o preenchimento. Não editar.",
            "",
        ])
    else:
        instrucoes.extend([
            "  5. A ECD atual não tem Bloco C (C050/C155) preenchido, então o",
            "     plano antigo não pôde ser recuperado automaticamente. Para",
            "     concluir a reconciliação, consulte diretamente o arquivo ECD",
            "     do ano anterior ou o razão do cliente.",
            "",
        ])
    instrucoes.extend([
        "Importação do arquivo preenchido:",
        "   Fora do escopo atual (§16.6). Em versões futuras, o arquivo",
        "   preenchido poderá ser reimportado via comando próprio para ativar",
        "   o modo integral nos cruzamentos que estavam abortados.",
        "",
        "Base legal:",
        "   IN RFB 2.003/2021 (ECD); ADE Cofis 01/2026 (Leiaute 9 — Bloco C).",
    ])

    for idx, linha in enumerate(instrucoes, start=5):
        ws.cell(row=idx, column=1, value=linha.format(cnpj=cnpj, ano=ano)).font = _FONT_INSTRUCAO

    ws.column_dimensions["A"].width = 90


def _aba_plano_contas(wb: Workbook, i050: list[dict]) -> None:
    ws = wb.create_sheet("Plano de Contas (atual)")

    colunas = [
        "COD_CTA (atual)", "NOME_CTA (atual)", "COD_NAT", "Natureza",
        "IND_CTA", "Nível", "COD_CTA_SUP",
        "COD_CTA antigo", "NOME antigo", "Observações",
    ]

    # Cabeçalho
    for col_idx, titulo in enumerate(colunas, start=1):
        cel = ws.cell(row=1, column=col_idx, value=titulo)
        cel.font = _FONT_HEADER
        # Últimas 3 colunas (preencher manualmente) em amarelo para destacar
        cel.fill = _FILL_AMARELO if col_idx >= 8 else _FILL_TEAL
        if col_idx >= 8:
            cel.font = Font(bold=True, color=_CINZA, name="Calibri", size=11)
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    # Ordena por COD_CTA para facilitar visualização
    i050_ord = sorted(i050, key=lambda r: (r.get("cod_cta") or ""))

    for row_idx, reg in enumerate(i050_ord, start=2):
        cod_nat = (reg.get("cod_nat") or "").strip()
        valores = [
            reg.get("cod_cta", ""),
            reg.get("cta", ""),
            cod_nat,
            _ROTULO_NATUREZA.get(cod_nat, ""),
            reg.get("ind_cta", ""),
            reg.get("nivel", ""),
            reg.get("cod_cta_sup", "") or "",
            "",  # COD_CTA antigo — auditor preenche
            "",  # NOME antigo — auditor preenche
            "",  # Observações — auditor preenche
        ]
        for col_idx, val in enumerate(valores, start=1):
            cel = ws.cell(row=row_idx, column=col_idx, value=val)
            cel.font = _FONT_NORMAL
            if col_idx >= 8:
                cel.fill = _FILL_AMARELO

    ws.freeze_panes = "A2"

    larguras = {1: 18, 2: 40, 3: 10, 4: 20, 5: 10, 6: 8, 7: 18, 8: 18, 9: 40, 10: 30}
    for col_idx, w in larguras.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def _aba_plano_antigo(wb: Workbook, c050: list[dict]) -> None:
    """Aba de consulta com o plano recuperado do exercício anterior (C050)."""
    ws = wb.create_sheet("Plano Antigo (C050)")

    colunas = [
        "COD_CTA (antigo)", "NOME_CTA (antigo)", "COD_NAT", "Natureza",
        "IND_CTA", "Nível", "COD_CTA_SUP",
    ]

    for col_idx, titulo in enumerate(colunas, start=1):
        cel = ws.cell(row=1, column=col_idx, value=titulo)
        cel.font = _FONT_HEADER
        cel.fill = _FILL_CINZA
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    c050_ord = sorted(c050, key=lambda r: (r.get("cod_cta") or ""))

    for row_idx, reg in enumerate(c050_ord, start=2):
        cod_nat = (reg.get("cod_nat") or "").strip()
        valores = [
            reg.get("cod_cta", ""),
            reg.get("cta", ""),
            cod_nat,
            _ROTULO_NATUREZA.get(cod_nat, ""),
            reg.get("ind_cta", ""),
            reg.get("nivel", ""),
            reg.get("cod_cta_sup", "") or "",
        ]
        for col_idx, val in enumerate(valores, start=1):
            cel = ws.cell(row=row_idx, column=col_idx, value=val)
            cel.font = _FONT_NORMAL

    ws.freeze_panes = "A2"

    larguras = {1: 18, 2: 40, 3: 10, 4: 20, 5: 10, 6: 8, 7: 18}
    for col_idx, w in larguras.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

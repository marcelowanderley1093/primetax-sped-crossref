"""
Geração do relatório Excel de diagnóstico (primetax-sped diagnose).

Estrutura de abas:
  1. "Oportunidades"       — cruzamentos da Camada 2 com impacto monetário.
  2. "Divergências"        — cruzamentos da Camada 1 (integridade).
  3. "Qualidade da Análise" — quatro tabelas CLAUDE.md §18.5:
       (a) Qualidade de reconciliação por SPED.
       (b) Cruzamentos em modo degradado.
       (c) Pendências recuperáveis (disponibilidade = pendente).
       (d) Limitações estruturais (disponibilidade = estruturalmente_ausente).

Identidade visual Primetax (CLAUDE.md §3):
  Cinza escuro: #53565A  /  Teal: #008C95
"""

import json
import re
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.db.repo import Repositorio

# --- Cores Primetax ---
_TEAL = "008C95"
_CINZA = "53565A"
_BRANCO = "FFFFFF"
_CINZA_CLARO = "F5F5F5"

_FILL_TEAL = PatternFill("solid", fgColor=_TEAL)
_FILL_CINZA = PatternFill("solid", fgColor=_CINZA)
_FILL_CINZA_CLARO = PatternFill("solid", fgColor=_CINZA_CLARO)

_FONT_HEADER = Font(bold=True, color=_BRANCO, name="Calibri", size=11)
_FONT_TITULO = Font(bold=True, color=_CINZA, name="Calibri", size=13)
_FONT_NORMAL = Font(name="Calibri", size=10)
_FONT_MONO = Font(name="Courier New", size=9)


def _cabecalho_aba(ws, colunas: list[str], fill=None) -> None:
    fill = fill or _FILL_TEAL
    for col_idx, titulo in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font = _FONT_HEADER
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30


def _ajustar_colunas(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def _moeda(v) -> str:
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(v)


def _truncar_json(evidencia_json: str, max_chars: int = 200) -> str:
    try:
        data = json.loads(evidencia_json)
        resumo = json.dumps(data, ensure_ascii=False)
        if len(resumo) > max_chars:
            return resumo[:max_chars] + "…"
        return resumo
    except Exception:
        return str(evidencia_json)[:max_chars]


def gerar(
    repo: Repositorio,
    ano_calendario: int,
    destino: Path,
) -> Path:
    """
    Gera o Excel de diagnóstico para CNPJ × ano-calendário.

    Args:
        repo: Repositório do cliente.
        ano_calendario: Ano-calendário do diagnóstico.
        destino: Caminho de saída do arquivo .xlsx.

    Returns:
        O caminho do arquivo gerado.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove aba padrão

    conn = repo.conexao()
    try:
        oportunidades = repo.consultar_oportunidades(conn, repo.cnpj, ano_calendario)
        divergencias = repo.consultar_divergencias(conn, repo.cnpj, ano_calendario)
        ctx = repo.consultar_sped_contexto(conn)
        meses = repo.consultar_meses_importados(conn, repo.cnpj, ano_calendario)
    finally:
        conn.close()

    _aba_oportunidades(wb, oportunidades)
    _aba_divergencias(wb, divergencias)
    _aba_qualidade(
        wb, ctx, meses, repo.cnpj, ano_calendario,
        oportunidades=oportunidades, divergencias=divergencias,
    )

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(destino))
    return destino


# ------------------------------------------------------------------
# Aba 1 — Oportunidades
# ------------------------------------------------------------------

def _aba_oportunidades(wb: Workbook, oportunidades: list[dict]) -> None:
    ws = wb.create_sheet("Oportunidades")
    colunas = [
        "Regra", "Descrição", "Severidade",
        "Impacto Conservador", "Impacto Máximo",
        "CNPJ", "Ano/Mês",
        "Arquivo Origem", "Linha", "Campos-Chave",
        "Gerado em",
    ]
    _cabecalho_aba(ws, colunas)

    total_cons = Decimal("0")
    total_max = Decimal("0")

    for row_idx, op in enumerate(oportunidades, start=2):
        evidencia = {}
        try:
            evs = json.loads(op.get("evidencia_json", "[]"))
            if evs:
                evidencia = evs[0] if isinstance(evs, list) else evs
        except Exception:
            pass

        campos_chave = evidencia.get("campos_chave", {})
        cons = Decimal(str(op.get("valor_impacto_conservador", 0)))
        maximo = Decimal(str(op.get("valor_impacto_maximo", 0)))
        total_cons += cons
        total_max += maximo

        valores = [
            op.get("codigo_regra", ""),
            op.get("descricao", ""),
            op.get("severidade", ""),
            float(cons),
            float(maximo),
            op.get("cnpj_declarante", ""),
            op.get("ano_mes", ""),
            evidencia.get("arquivo_origem", ""),
            evidencia.get("linha_arquivo", ""),
            json.dumps(campos_chave, ensure_ascii=False, default=str)[:300],
            op.get("gerado_em", ""),
        ]
        for col_idx, val in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _FONT_NORMAL
            if col_idx in (4, 5):
                cell.number_format = 'R$ #,##0.00'
            if row_idx % 2 == 0:
                cell.fill = _FILL_CINZA_CLARO

    # Linha de totais
    if oportunidades:
        row_total = len(oportunidades) + 2
        ws.cell(row=row_total, column=3, value="TOTAL").font = Font(bold=True)
        tc = ws.cell(row=row_total, column=4, value=float(total_cons))
        tc.font = Font(bold=True)
        tc.number_format = 'R$ #,##0.00'
        tm = ws.cell(row=row_total, column=5, value=float(total_max))
        tm.font = Font(bold=True)
        tm.number_format = 'R$ #,##0.00'

    ws.freeze_panes = "A2"
    _ajustar_colunas(ws)


# ------------------------------------------------------------------
# Aba 2 — Divergências
# ------------------------------------------------------------------

def _aba_divergencias(wb: Workbook, divergencias: list[dict]) -> None:
    ws = wb.create_sheet("Divergências de Integridade")
    colunas = ["Regra", "Descrição", "Severidade", "CNPJ", "Ano/Mês", "Evidência", "Gerado em"]
    _cabecalho_aba(ws, colunas, fill=_FILL_CINZA)

    for row_idx, div in enumerate(divergencias, start=2):
        valores = [
            div.get("codigo_regra", ""),
            div.get("descricao", ""),
            div.get("severidade", ""),
            div.get("cnpj_declarante", ""),
            div.get("ano_mes", ""),
            _truncar_json(div.get("evidencia_json", "")),
            div.get("gerado_em", ""),
        ]
        for col_idx, val in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _FONT_NORMAL
            if row_idx % 2 == 0:
                cell.fill = _FILL_CINZA_CLARO

    ws.freeze_panes = "A2"
    _ajustar_colunas(ws)


# ------------------------------------------------------------------
# Aba 3 — Qualidade da Análise
# ------------------------------------------------------------------

# ------------------------------------------------------------------
# Metadados dos cruzamentos (CLAUDE.md §8)
# Mapa codigo_regra → (dependências SPED, descrição curta)
# Usado para gerar as tabelas 3 e 4 da aba Qualidade da Análise.
# ------------------------------------------------------------------

_METADATA_REGRAS: list[tuple[str, list[str], str]] = [
    # Camada 1 — Integridade
    ("CR-01", ["efd_contribuicoes"], "Integridade 9900/9999"),
    ("CR-02", ["efd_contribuicoes"], "Unicidade M200/M600"),
    ("CR-03", ["efd_contribuicoes"], "Presença e validade do 0110"),
    ("CR-04", ["efd_contribuicoes"], "Consistência 0110 × 0111 (rateio/apropriação)"),
    ("CR-05", ["efd_contribuicoes"], "Hierarquia pai-filho dos registros"),
    ("CR-06", ["efd_contribuicoes"], "Coerência CNPJ/período entre SPEDs"),
    # Camada 2 — EFD-Contribuições
    ("CR-07", ["efd_contribuicoes"], "Tese 69 — exclusão ICMS da base em C170"),
    ("CR-08", ["efd_contribuicoes"], "Tese 69 em C181/C185 (NFC-e)"),
    ("CR-09", ["efd_contribuicoes"], "Tese 69 em D100/D200 (transporte/comunicação)"),
    ("CR-10", ["efd_contribuicoes"], "CST × CFOP × alíquota em C170"),
    ("CR-11", ["efd_contribuicoes"], "TIPO_ITEM 07 com CFOP de insumo (REsp 1.221.170)"),
    ("CR-12", ["efd_contribuicoes"], "CST 70-75 — reclassificação"),
    ("CR-13", ["efd_contribuicoes"], "Ausência de F100 para operações habituais"),
    ("CR-14", ["efd_contribuicoes"], "F120 — crédito sobre depreciação"),
    ("CR-15", ["efd_contribuicoes"], "F130 — crédito sobre aquisição de imobilizado"),
    ("CR-16", ["efd_contribuicoes"], "Retenções PIS na fonte (F600)"),
    ("CR-17", ["efd_contribuicoes"], "Retenções COFINS na fonte (F700)"),
    ("CR-18", ["efd_contribuicoes"], "Retenções F800 (eventos corporativos/CPRB)"),
    ("CR-19", ["efd_contribuicoes"], "Retenções prescritas em 1300/1700"),
    ("CR-20", ["efd_contribuicoes"], "Créditos presumidos setoriais"),
    ("CR-21", ["efd_contribuicoes"], "Transporte de carga subcontratado"),
    ("CR-22", ["efd_contribuicoes"], "F150 — estoque de abertura"),
    ("CR-23", ["efd_contribuicoes"], "CST 98/99 em C170 para imobilizado"),
    ("CR-25", ["efd_contribuicoes"], "Saldos em 1100/1500 (carry-forward)"),
    ("CR-26", ["efd_contribuicoes"], "Ajuste da base via M215/M615 (Tese 69 pós-2019)"),
    ("CR-27", ["efd_contribuicoes"], "Créditos vedados por área (F120/F130)"),
    ("CR-28", ["efd_contribuicoes"], "Rateio proporcional × apropriação direta"),
    # Camada 3 — Consistência intra-EFD-Contribuições
    ("CR-29", ["efd_contribuicoes"], "M210 ↔ M100 ↔ M200 (fluxo PIS)"),
    ("CR-30", ["efd_contribuicoes"], "M610 ↔ M500 ↔ M600 (fluxo COFINS)"),
    ("CR-31", ["efd_contribuicoes"], "Σ bases itens ↔ M210 (PIS)"),
    ("CR-32", ["efd_contribuicoes"], "Σ bases créditos ↔ M105 (PIS)"),
    ("CR-33", ["efd_contribuicoes"], "Σ bases itens ↔ M610 (COFINS)"),
    ("CR-34", ["efd_contribuicoes"], "Σ bases créditos ↔ M505 (COFINS)"),
    # Sprint 6 — inter-SPED EFD ICMS/IPI
    ("CR-35", ["efd_contribuicoes", "efd_icms"], "CIAP (G125) × F120/F130"),
    ("CR-36", ["efd_contribuicoes", "efd_icms"], "H010 inventário × F150 estoque abertura"),
    ("CR-37", ["efd_contribuicoes", "efd_icms"], "CFOP de exportação × CST de receita"),
    # Sprint 7 — inter-SPED ECD
    ("CR-38", ["efd_contribuicoes", "ecd"], "I155 × M200/M600 (receita × base)"),
    ("CR-39", ["efd_contribuicoes", "ecd"], "J150 DRE × apuração anual EFD"),
    ("CR-40", ["efd_contribuicoes", "ecd"], "COD_CTA em C170/F100 × plano I050"),
    ("CR-41", ["efd_contribuicoes", "ecd"], "Lançamento extemporâneo I200 × 1101/1501"),
    ("CR-42", ["efd_contribuicoes", "ecd"], "TIPO_ITEM em 0200 × conta I050"),
    # Sprint 8 — ECF × ECD / ECF isolada
    ("CR-43", ["ecf", "ecd"], "K155/K355 × I155 (saldos contábeis)"),
    ("CR-44", ["ecf", "ecd"], "M300/M312 × I200 (NUM_LCTO)"),
    ("CR-45", ["ecf"], "M500 Parte B estagnada"),
    ("CR-46", ["ecf"], "X480 × M300 (benefício não aproveitado)"),
    ("CR-47", ["ecf"], "Y570 IRRF/CSRF retido não compensado"),
    # §8.6 — cruzamentos adicionais
    ("CR-48", ["ecf"], "9100 avisos do PGE na transmissão"),
    ("CR-49", ["ecf"], "X460 Lei do Bem — dispêndios P&D sem exclusão aproveitada"),
]

# Mapa SPED interno → campo em _sped_contexto
_CAMPO_POR_SPED = {
    "efd_contribuicoes": "disponibilidade_efd_contrib",
    "efd_icms": "disponibilidade_efd_icms",
    "ecd": "disponibilidade_ecd",
    "ecf": "disponibilidade_ecf",
    "bloco_i": "disponibilidade_bloco_i",
}

# Rótulo amigável para cada SPED
_ROTULO_SPED = {
    "efd_contribuicoes": "EFD-Contribuições",
    "efd_icms": "EFD ICMS/IPI",
    "ecd": "ECD",
    "ecf": "ECF",
    "bloco_i": "Bloco I (financeiras)",
}

# Base legal e causa declarada para cada ausência estrutural (CLAUDE.md §18.8)
_AUSENCIA_ESTRUTURAL = {
    "efd_icms": (
        "PJ sem operações de circulação de mercadoria (atividade ISS)",
        "Ajuste SINIEF 02/2009, cláusula 5ª",
    ),
    "ecd": (
        "PJ adotou Livro Caixa (ECF.0010.TIP_ESC_PRE='L')",
        "Art. 45 parágrafo único Lei 8.981/1995; IN RFB 2.003/2021",
    ),
    "ecf": (
        "PJ optante do Simples Nacional ou inativa",
        "IN RFB 2.004/2021",
    ),
    "bloco_i": (
        "Bloco I aplicável apenas a instituições financeiras/seguradoras",
        "IN RFB 1.252/2012",
    ),
}


def _disponibilidade(ctx: dict | None, sped: str) -> str:
    """Retorna o estado de disponibilidade de um SPED a partir do contexto."""
    if ctx is None:
        return "pendente"
    campo = _CAMPO_POR_SPED.get(sped)
    if campo is None:
        return "pendente"
    return ctx.get(campo) or "pendente"


def _estado_efetivo(ctx: dict | None, dependencias: list[str]) -> str:
    """Estado efetivo do cruzamento (§18.3): estrutural > pendente > importada."""
    estados = [_disponibilidade(ctx, d) for d in dependencias]
    if "estruturalmente_ausente" in estados:
        return "estruturalmente_ausente"
    if "pendente" in estados:
        return "pendente"
    return "importada"


def _cabecalho_secao(ws, row: int, titulo: str) -> int:
    c = ws.cell(row=row, column=1, value=titulo)
    c.font = Font(bold=True, color=_CINZA, name="Calibri", size=12)
    return row + 1


def _cabecalho_tabela(ws, row: int, colunas: list[str], fill=None) -> int:
    fill = fill or _FILL_TEAL
    for col_idx, h in enumerate(colunas, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = _FONT_HEADER
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return row + 1


_GRANULARIDADE_DEGRADADA = {
    "CR-38": "COD_NAT='04' (plano combinado I050 ∪ C050)",
    "CR-39": "IND_GRP_DRE (natureza da DRE)",
    "CR-43": "COD_NAT agregado (saldos patrimoniais e de resultado)",
}


def _extrair_modo_degradado(achados: list[dict]) -> list[tuple[str, str, str]]:
    """
    Varre oportunidades + divergências e retorna triplas
    (codigo_regra, granularidade, motivo) para entradas com modo_execucao='degradado'.

    A granularidade efetiva vem de `_GRANULARIDADE_DEGRADADA[codigo]`, que documenta
    exatamente qual eixo de agregação cada cruzamento usa em modo degradado.
    """
    encontrados: list[tuple[str, str, str]] = []
    vistos: set[str] = set()
    for a in achados:
        try:
            evs = json.loads(a.get("evidencia_json", "[]"))
        except Exception:
            continue
        if not evs:
            continue
        ev = evs[0] if isinstance(evs, list) else evs
        ck = ev.get("campos_chave", {}) if isinstance(ev, dict) else {}
        if ck.get("modo_execucao") != "degradado":
            continue
        codigo = a.get("codigo_regra", "")
        if codigo in vistos:
            continue
        vistos.add(codigo)
        granularidade = _GRANULARIDADE_DEGRADADA.get(codigo, "COD_NAT agregado")
        motivo = (
            f"ECD com reconciliação '{ck.get('reconciliacao_plano_contas', '?')}'"
            f" — Bloco C ausente ou parcial"
        )
        encontrados.append((codigo, granularidade, motivo))
    return encontrados


def _aba_qualidade(
    wb: Workbook,
    ctx: dict | None,
    meses: list[int],
    cnpj: str,
    ano_calendario: int,
    *,
    oportunidades: list[dict] | None = None,
    divergencias: list[dict] | None = None,
) -> None:
    ws = wb.create_sheet("Qualidade da Análise")
    row = 1

    ws.cell(row=row, column=1, value=f"Qualidade da Análise — CNPJ {cnpj} / AC {ano_calendario}")
    ws.cell(row=row, column=1).font = _FONT_TITULO
    row += 2

    meses_str = ", ".join(str(m) for m in meses) if meses else "—"

    # ---------- Tabela 1 — Qualidade por SPED ----------
    row = _cabecalho_secao(ws, row, "1. Qualidade de SPEDs importados (§16.5 / §18.5)")
    row = _cabecalho_tabela(
        ws, row,
        ["SPED", "Disponibilidade", "Reconciliação plano de contas", "Observação"],
    )

    reconc_geral = (ctx or {}).get("reconciliacao_plano_contas")
    for sped_interno in ("efd_contribuicoes", "efd_icms", "ecd", "ecf", "bloco_i"):
        rotulo = _ROTULO_SPED[sped_interno]
        estado = _disponibilidade(ctx, sped_interno)

        reconc = "—"
        obs = ""
        if estado == "pendente":
            obs = "Arquivo não importado. Cruzamentos dependentes ficam em Pendências (tabela 3)."
        elif estado == "estruturalmente_ausente":
            causa, _base = _AUSENCIA_ESTRUTURAL.get(sped_interno, ("", ""))
            obs = f"Ausência estrutural confirmada: {causa}"
        elif estado == "importada":
            if sped_interno == "efd_contribuicoes":
                obs = f"{len(meses)} competência(s) mensal(is): {meses_str}"
            elif sped_interno == "efd_icms":
                obs = f"{len(meses)} competência(s) mensal(is): {meses_str}"
            elif sped_interno == "ecd":
                obs = f"Exercício {ano_calendario}"
                reconc = reconc_geral or "não avaliada"
            elif sped_interno == "ecf":
                obs = f"Ano-calendário {ano_calendario}"
            elif sped_interno == "bloco_i":
                obs = "Processado dentro da EFD-Contribuições"

        for col_idx, val in enumerate([rotulo, estado, reconc, obs], start=1):
            cel = ws.cell(row=row, column=col_idx, value=val)
            cel.font = _FONT_NORMAL
        row += 1
    row += 2

    # ---------- Tabela 2 — Cruzamentos em modo degradado ----------
    row = _cabecalho_secao(ws, row, "2. Cruzamentos executados em modo degradado (§16.5)")
    row = _cabecalho_tabela(
        ws, row,
        ["Cruzamento", "Descrição", "Granularidade efetiva", "Motivo"],
        fill=_FILL_CINZA,
    )

    # Modo degradado = achados com evidencia.campos_chave.modo_execucao='degradado'.
    # Iteração dinâmica sobre oportunidades e divergências registradas no banco.
    degradados = _extrair_modo_degradado((oportunidades or []) + (divergencias or []))
    descricoes_por_regra = {codigo: descricao for codigo, _deps, descricao in _METADATA_REGRAS}

    if degradados:
        for codigo, granularidade, motivo in degradados:
            ws.cell(row=row, column=1, value=codigo).font = _FONT_NORMAL
            ws.cell(
                row=row, column=2,
                value=descricoes_por_regra.get(codigo, ""),
            ).font = _FONT_NORMAL
            ws.cell(row=row, column=3, value=granularidade).font = _FONT_NORMAL
            ws.cell(row=row, column=4, value=motivo).font = _FONT_NORMAL
            row += 1
    else:
        ws.cell(row=row, column=1, value="(nenhum)").font = _FONT_NORMAL
        ws.cell(
            row=row, column=2,
            value="Todos os cruzamentos executados em modo integral.",
        ).font = _FONT_NORMAL
        row += 1
    row += 2

    # ---------- Tabela 3 — Pendências recuperáveis ----------
    row = _cabecalho_secao(ws, row, "3. Pendências recuperáveis (ação do auditor requerida) (§18.5)")
    row = _cabecalho_tabela(
        ws, row,
        ["Cruzamento", "Descrição", "Dependência faltante", "Ação sugerida"],
        fill=_FILL_CINZA,
    )

    pendentes_encontrados = False
    for codigo, deps, descricao in _METADATA_REGRAS:
        if _estado_efetivo(ctx, deps) != "pendente":
            continue
        faltantes = [
            _ROTULO_SPED[d] for d in deps
            if _disponibilidade(ctx, d) == "pendente"
        ]
        acao = (
            f"Importar {', '.join(faltantes)} para CNPJ {cnpj} / AC {ano_calendario}"
        )
        for col_idx, val in enumerate(
            [codigo, descricao, ", ".join(faltantes), acao], start=1
        ):
            cel = ws.cell(row=row, column=col_idx, value=val)
            cel.font = _FONT_NORMAL
        row += 1
        pendentes_encontrados = True

    if not pendentes_encontrados:
        ws.cell(row=row, column=1, value="(nenhuma)").font = _FONT_NORMAL
        ws.cell(
            row=row, column=2,
            value="Diagnóstico completo para os SPEDs disponíveis.",
        ).font = _FONT_NORMAL
        row += 1
    row += 2

    # ---------- Tabela 4 — Limitações estruturais ----------
    row = _cabecalho_secao(
        ws, row,
        "4. Limitações estruturais (inaplicáveis por opção fiscal da PJ) (§18.5)",
    )
    row = _cabecalho_tabela(
        ws, row,
        ["Cruzamento", "Descrição", "Causa declarada", "Base legal"],
        fill=_FILL_CINZA,
    )

    limitacoes_encontradas = False
    for codigo, deps, descricao in _METADATA_REGRAS:
        if _estado_efetivo(ctx, deps) != "estruturalmente_ausente":
            continue
        # Identifica o SPED que causou a ausência estrutural
        sped_ausente = next(
            (d for d in deps if _disponibilidade(ctx, d) == "estruturalmente_ausente"),
            None,
        )
        causa, base_legal = _AUSENCIA_ESTRUTURAL.get(
            sped_ausente or "", ("Ausência estrutural", "Ver CLAUDE.md §18.8")
        )
        for col_idx, val in enumerate(
            [codigo, descricao, causa, base_legal], start=1
        ):
            cel = ws.cell(row=row, column=col_idx, value=val)
            cel.font = _FONT_NORMAL
        row += 1
        limitacoes_encontradas = True

    if not limitacoes_encontradas:
        ws.cell(row=row, column=1, value="(nenhuma)").font = _FONT_NORMAL
        ws.cell(
            row=row, column=2,
            value="Nenhum SPED foi classificado como estruturalmente ausente.",
        ).font = _FONT_NORMAL
        row += 1

    _ajustar_colunas(ws)

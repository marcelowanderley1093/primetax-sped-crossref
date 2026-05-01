"""
Testes do DataTable + MoneyCellDelegate (Bloco 5 §C1, §C2).

Cobertura desta iteração:
  - construção, set_rows, contadores
  - filtro reduz visíveis e limpa restaura
  - ordenação por header (proxy)
  - selected_row e visible_rows
  - signal row_activated em duplo-clique
  - exportação Excel grava arquivo válido
  - empty state quando sem linhas
  - Money.format() em casos comuns (positivo, negativo, zero)
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from PySide6.QtCore import QModelIndex, Qt

from src.gui.widgets import (
    BadgeStatus,
    ColumnSpec,
    DataTable,
    Money,
    MoneyOptions,
)


def _rows_amostra() -> list[dict]:
    return [
        {
            "regra": "CR-07",
            "descricao": "Tese 69 em C170",
            "severidade": BadgeStatus.ALTO,
            "impacto": Decimal("523000.00"),
            "evidencia": {"arquivo": "/p/efd_c.txt", "linha": 1847,
                          "bloco": "C", "registro": "C170"},
        },
        {
            "regra": "CR-19",
            "descricao": "Retenções prescritas em 1300",
            "severidade": BadgeStatus.ALTO,
            "impacto": Decimal("45000.00"),
            "evidencia": {"arquivo": "/p/efd_c.txt", "linha": 9120,
                          "bloco": "1", "registro": "1300"},
        },
        {
            "regra": "CR-22",
            "descricao": "F150 estoque de abertura",
            "severidade": BadgeStatus.MEDIO,
            "impacto": Decimal("0"),
            "evidencia": None,
        },
    ]


def _columns_amostra() -> list[ColumnSpec]:
    return [
        ColumnSpec(id="regra", header="Regra", kind="text", width=80),
        ColumnSpec(id="descricao", header="Descrição", kind="text", width=240),
        ColumnSpec(id="severidade", header="Severidade", kind="badge", width=110),
        ColumnSpec(id="impacto", header="Impacto", kind="money", width=140),
        ColumnSpec(id="evidencia", header="Linha", kind="code_link", width=70),
    ]


# --------------------------------------------------------------------
# Money helper
# --------------------------------------------------------------------

class TestMoney:
    def test_formato_positivo(self):
        assert Money.format(Decimal("1234567.89")) == "R$ 1.234.567,89"

    def test_formato_negativo_parens(self):
        assert Money.format(Decimal("-1234.56")) == "(R$ 1.234,56)"

    def test_formato_negativo_minus(self):
        opts = MoneyOptions(negative_style="minus")
        assert Money.format(Decimal("-1234.56"), opts) == "-R$ 1.234,56"

    def test_zero_show_zero_true(self):
        assert Money.format(Decimal("0")) == "R$ 0,00"

    def test_valor_invalido(self):
        assert Money.format("texto-quebrado") == "—"

    def test_accessible(self):
        assert "reais" in Money.format_accessible(Decimal("1234.56"))


# --------------------------------------------------------------------
# DataTable
# --------------------------------------------------------------------

class TestDataTable:
    def test_construcao_sem_linhas(self, qtbot):
        dt = DataTable(_columns_amostra(), [])
        qtbot.addWidget(dt)
        assert dt.total_rows() == 0
        assert dt.visible_count() == 0

    def test_construcao_com_3_linhas(self, qtbot):
        dt = DataTable(_columns_amostra(), _rows_amostra())
        qtbot.addWidget(dt)
        assert dt.total_rows() == 3
        assert dt.visible_count() == 3

    def test_set_rows_substitui_completamente(self, qtbot):
        dt = DataTable(_columns_amostra(), _rows_amostra())
        qtbot.addWidget(dt)
        nova = [_rows_amostra()[0]]
        dt.set_rows(nova)
        assert dt.total_rows() == 1

    def test_filtro_reduz_visiveis(self, qtbot):
        dt = DataTable(_columns_amostra(), _rows_amostra())
        qtbot.addWidget(dt)
        dt.set_filter("Tese 69")
        assert dt.visible_count() == 1
        rows = dt.visible_rows()
        assert rows[0]["regra"] == "CR-07"

    def test_filtro_vazio_restaura_todas(self, qtbot):
        dt = DataTable(_columns_amostra(), _rows_amostra())
        qtbot.addWidget(dt)
        dt.set_filter("Tese 69")
        dt.set_filter("")
        assert dt.visible_count() == 3

    def test_filtro_case_insensitive(self, qtbot):
        dt = DataTable(_columns_amostra(), _rows_amostra())
        qtbot.addWidget(dt)
        dt.set_filter("TESE")
        assert dt.visible_count() == 1
        dt.set_filter("tese")
        assert dt.visible_count() == 1

    def test_visible_rows_respeita_filtro(self, qtbot):
        dt = DataTable(_columns_amostra(), _rows_amostra())
        qtbot.addWidget(dt)
        dt.set_filter("CR-19")
        rows = dt.visible_rows()
        assert len(rows) == 1
        assert rows[0]["regra"] == "CR-19"

    def test_add_row_aumenta_total(self, qtbot):
        dt = DataTable(_columns_amostra(), _rows_amostra())
        qtbot.addWidget(dt)
        dt.add_row({
            "regra": "CR-46",
            "descricao": "X480 sem M300",
            "severidade": BadgeStatus.MEDIO,
            "impacto": Decimal("12000"),
            "evidencia": None,
        })
        assert dt.total_rows() == 4

    def test_remove_rows_via_predicate(self, qtbot):
        dt = DataTable(_columns_amostra(), _rows_amostra())
        qtbot.addWidget(dt)
        n = dt.remove_rows(lambda r: r["regra"] == "CR-19")
        assert n == 1
        assert dt.total_rows() == 2

    def test_update_row_via_predicate(self, qtbot):
        dt = DataTable(_columns_amostra(), _rows_amostra())
        qtbot.addWidget(dt)
        n = dt.update_row(
            lambda r: r["regra"] == "CR-22",
            {"impacto": Decimal("99000")},
        )
        assert n == 1
        encontrada = next(r for r in dt._model.rows() if r["regra"] == "CR-22")
        assert encontrada["impacto"] == Decimal("99000")

    def test_selected_row_inicial_eh_none(self, qtbot):
        dt = DataTable(_columns_amostra(), _rows_amostra())
        qtbot.addWidget(dt)
        assert dt.selected_row() is None

    def test_selecionar_primeira_linha_emite_signal(self, qtbot):
        dt = DataTable(_columns_amostra(), _rows_amostra())
        qtbot.addWidget(dt)
        with qtbot.waitSignal(dt.row_selected, timeout=500) as blocker:
            idx = dt._proxy.index(0, 0)
            dt._view.selectionModel().setCurrentIndex(
                idx, dt._view.selectionModel().SelectionFlag.SelectCurrent
                | dt._view.selectionModel().SelectionFlag.Rows
            )
        assert blocker.args[0]["regra"] in ("CR-07", "CR-19", "CR-22")

    def test_export_emite_visible_rows(self, qtbot):
        dt = DataTable(_columns_amostra(), _rows_amostra())
        qtbot.addWidget(dt)
        dt.set_filter("Tese 69")
        with qtbot.waitSignal(dt.export_requested, timeout=500) as blocker:
            dt._on_export_clicked()
        rows = blocker.args[0]
        assert len(rows) == 1
        assert rows[0]["regra"] == "CR-07"

    def test_export_xlsx_grava_arquivo_valido(self, qtbot, tmp_path):
        dt = DataTable(_columns_amostra(), _rows_amostra())
        qtbot.addWidget(dt)
        destino = tmp_path / "export.xlsx"
        dt._gerar_xlsx(destino, dt.visible_rows())

        assert destino.exists()
        # Valida estrutura
        from openpyxl import load_workbook
        wb = load_workbook(destino)
        ws = wb.active
        # Header
        assert ws.cell(1, 1).value == "Regra"
        assert ws.cell(1, 4).value == "Impacto"
        # Primeira linha de dados
        regras = {ws.cell(r, 1).value for r in (2, 3, 4)}
        assert regras == {"CR-07", "CR-19", "CR-22"}

    def test_ordenacao_por_impacto(self, qtbot):
        dt = DataTable(_columns_amostra(), _rows_amostra())
        qtbot.addWidget(dt)
        # Ordena coluna "Impacto" desc
        dt._view.sortByColumn(3, Qt.SortOrder.DescendingOrder)
        rows = dt.visible_rows()
        assert rows[0]["regra"] == "CR-07"  # 523000 maior
        assert rows[-1]["regra"] == "CR-22"  # 0 menor

    def test_empty_state_quando_filtro_zera_visiveis(self, qtbot):
        dt = DataTable(_columns_amostra(), _rows_amostra())
        qtbot.addWidget(dt)
        dt.show()
        dt.set_filter("inexistente-zzz")
        assert dt.visible_count() == 0
        assert not dt._empty_label.isHidden()

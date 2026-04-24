"""
Testes do importador de template de reconciliação manual (CLAUDE.md §16.6).

Cobertura:
  - Arquivo inválido/inexistente levanta ValueError.
  - Template gerado sem nenhum mapeamento preenchido levanta ValueError.
  - Template com mapeamentos válidos insere overrides no banco.
  - Classificação da reconciliação é elevada a 'integra' quando cobertura
    dos overrides atinge 50% das contas analíticas do I050.
  - resolver_cod_cta() usa os overrides corretamente.
  - Re-importação com substituir=True limpa overrides anteriores.
"""

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from src.crossref.common import reconciliacao_plano_contas
from src.db.repo import Repositorio
from src.parsers import ecd
from src.reports import excel_reconciliacao, excel_reconciliacao_import


def _importar_ecd(ecd_path: Path, tmp_path: Path, cnpj: str, ano: int):
    db_dir = tmp_path / "db"
    ecd.importar(
        ecd_path, encoding_override="utf8", prompt_operador=False, base_dir_db=db_dir
    )
    return Repositorio(cnpj, ano, base_dir=db_dir)


def _preencher_template(
    template_path: Path, saida_path: Path, mapeamentos: dict[str, str]
) -> None:
    """Carrega template gerado e preenche a coluna 'COD_CTA antigo' para cada mapeamento."""
    wb = load_workbook(str(template_path))
    ws = wb["Plano de Contas (atual)"]
    for row_idx in range(2, ws.max_row + 1):
        cod_atual = ws.cell(row=row_idx, column=1).value
        if cod_atual and cod_atual in mapeamentos:
            ws.cell(row=row_idx, column=8, value=mapeamentos[cod_atual])
            ws.cell(row=row_idx, column=9, value="")
    wb.save(str(saida_path))


class TestImportErros:
    def test_arquivo_inexistente_levanta_value_error(self, tmp_path):
        db_dir = tmp_path / "db"
        repo = Repositorio("00000000000100", 2025, base_dir=db_dir)
        repo.criar_banco()

        with pytest.raises(ValueError, match="não encontrado"):
            excel_reconciliacao_import.importar_template(
                repo, 2025, tmp_path / "inexistente.xlsx"
            )

    def test_arquivo_sem_aba_plano_levanta_value_error(self, tmp_path):
        arq = tmp_path / "errado.xlsx"
        wb = Workbook()
        wb.active.title = "Outra Aba"
        wb.save(str(arq))

        db_dir = tmp_path / "db"
        repo = Repositorio("00000000000100", 2025, base_dir=db_dir)
        repo.criar_banco()

        with pytest.raises(ValueError, match="'Plano de Contas \\(atual\\)' não encontrada"):
            excel_reconciliacao_import.importar_template(repo, 2025, arq)

    def test_template_sem_preenchimento_levanta_value_error(
        self, fixture_ecd_reconciliacao_mudanc_pc, tmp_path
    ):
        """Template gerado mas nenhuma linha com COD_CTA antigo preenchida."""
        repo = _importar_ecd(
            fixture_ecd_reconciliacao_mudanc_pc, tmp_path, "00000000000148", 2025
        )
        template = tmp_path / "tpl.xlsx"
        excel_reconciliacao.gerar(repo, 2025, template)

        with pytest.raises(ValueError, match="Nenhuma linha"):
            excel_reconciliacao_import.importar_template(repo, 2025, template)


class TestImportValido:
    def test_importa_mapeamentos_e_persiste(
        self, fixture_ecd_reconciliacao_mudanc_pc, tmp_path
    ):
        """Template preenchido com 2 mapeamentos → 2 overrides no banco."""
        repo = _importar_ecd(
            fixture_ecd_reconciliacao_mudanc_pc, tmp_path, "00000000000148", 2025
        )
        template = tmp_path / "tpl.xlsx"
        excel_reconciliacao.gerar(repo, 2025, template)

        preenchido = tmp_path / "preenchido.xlsx"
        _preencher_template(template, preenchido, {
            "1.01.001": "1001",
            "3.01.001": "3001",
        })

        resultado = excel_reconciliacao_import.importar_template(
            repo, 2025, preenchido
        )
        assert resultado.linhas_importadas == 2
        assert resultado.cnpj == "00000000000148"

        conn = repo.conexao()
        try:
            overrides = repo.consultar_reconciliacao_overrides(conn, "00000000000148", 2025)
        finally:
            conn.close()
        mapa = {o["cod_cta_atual"]: o["cod_cta_antigo"] for o in overrides}
        assert mapa == {"1.01.001": "1001", "3.01.001": "3001"}

    def test_overrides_elevam_classificacao_para_integra(
        self, fixture_ecd_reconciliacao_mudanc_pc, tmp_path
    ):
        """Cobertura de overrides ≥50% das contas analíticas → estado vira 'integra'."""
        repo = _importar_ecd(
            fixture_ecd_reconciliacao_mudanc_pc, tmp_path, "00000000000148", 2025
        )

        # Confirma estado inicial = ausente
        conn = repo.conexao()
        try:
            estado_inicial = reconciliacao_plano_contas.classificar_reconciliacao(
                repo, conn, "00000000000148", 2025
            )
        finally:
            conn.close()
        assert estado_inicial == "ausente"

        template = tmp_path / "tpl.xlsx"
        excel_reconciliacao.gerar(repo, 2025, template)
        preenchido = tmp_path / "preenchido.xlsx"
        _preencher_template(template, preenchido, {
            "1.01.001": "1001",
            "3.01.001": "3001",
        })
        excel_reconciliacao_import.importar_template(repo, 2025, preenchido)

        conn = repo.conexao()
        try:
            estado_final = reconciliacao_plano_contas.classificar_reconciliacao(
                repo, conn, "00000000000148", 2025
            )
        finally:
            conn.close()
        assert estado_final == "integra"

    def test_resolver_cod_cta_usa_overrides_bidirecional(
        self, fixture_ecd_reconciliacao_mudanc_pc, tmp_path
    ):
        """resolver_cod_cta: lookup por cod_cta_atual E por cod_cta_antigo."""
        repo = _importar_ecd(
            fixture_ecd_reconciliacao_mudanc_pc, tmp_path, "00000000000148", 2025
        )
        template = tmp_path / "tpl.xlsx"
        excel_reconciliacao.gerar(repo, 2025, template)
        preenchido = tmp_path / "preenchido.xlsx"
        _preencher_template(template, preenchido, {
            "1.01.001": "1001",
            "3.01.001": "3001",
        })
        excel_reconciliacao_import.importar_template(repo, 2025, preenchido)

        conn = repo.conexao()
        try:
            # Passagem pelo código atual → passthrough
            r1 = reconciliacao_plano_contas.resolver_cod_cta(
                repo, conn, "00000000000148", 2025, "1.01.001"
            )
            # Passagem pelo código antigo → retorna o atual
            r2 = reconciliacao_plano_contas.resolver_cod_cta(
                repo, conn, "00000000000148", 2025, "1001"
            )
            # Conta não mapeada → None (sem override + classificação virou integra
            # por cobertura global, então passthrough)
            r3 = reconciliacao_plano_contas.resolver_cod_cta(
                repo, conn, "00000000000148", 2025, "9.99.999"
            )
        finally:
            conn.close()

        assert r1 == "1.01.001"
        assert r2 == "1.01.001"
        assert r3 == "9.99.999"  # integra por cobertura → passthrough

    def test_reimportacao_substitui_overrides_anteriores(
        self, fixture_ecd_reconciliacao_mudanc_pc, tmp_path
    ):
        """Import com substituir=True (default) descarta mapeamentos anteriores."""
        repo = _importar_ecd(
            fixture_ecd_reconciliacao_mudanc_pc, tmp_path, "00000000000148", 2025
        )
        template = tmp_path / "tpl.xlsx"
        excel_reconciliacao.gerar(repo, 2025, template)

        preenchido1 = tmp_path / "p1.xlsx"
        _preencher_template(template, preenchido1, {"1.01.001": "1001"})
        excel_reconciliacao_import.importar_template(repo, 2025, preenchido1)

        preenchido2 = tmp_path / "p2.xlsx"
        _preencher_template(template, preenchido2, {"3.01.001": "3001"})
        excel_reconciliacao_import.importar_template(repo, 2025, preenchido2)

        conn = repo.conexao()
        try:
            overrides = repo.consultar_reconciliacao_overrides(conn, "00000000000148", 2025)
        finally:
            conn.close()
        # Deve ter apenas o último mapeamento (o primeiro foi limpo)
        assert len(overrides) == 1
        assert overrides[0]["cod_cta_atual"] == "3.01.001"
        assert overrides[0]["cod_cta_antigo"] == "3001"
